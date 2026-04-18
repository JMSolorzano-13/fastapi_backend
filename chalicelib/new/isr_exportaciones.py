from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from chalicelib.bus import get_global_bus
from chalicelib.controllers.common_utils.export_xlsx import query_to_xlsx
from chalicelib.new.cfdi_processor.domain.cfdi_exporter import CFDIExporter
from chalicelib.new.cfdi_processor.infra.cfdi_export_repository_sa import CFDIExportRepositorySA
from chalicelib.new.query.domain import DownloadType
from chalicelib.schema.models.tenant.cfdi_export import CfdiExport as ExportRequestORM

# Configuración de formato para filas totales ISR
ISR_CONCEPT_CONFIG = {
    "Gastos de nómina gravada": {"prefix": "(+) "},
    "Gastos de nómina exenta": {},
    "Gastos de nómina exenta deducible": {"prefix": "(+) ", "suffix_dynamic": "porcentaje"},
    "Gastos de nómina deducibles": {"prefix": "(=) ", "bold": True},
    "Compras y gastos facturas de contado": {"prefix": "(+) "},
    "Compras y gastos CFDI de pago": {"prefix": "(+) "},
    "Devoluciones, descuentos y bonificaciones facturadas": {"prefix": "(+) "},
    "Devoluciones, descuentos y bonificaciones en ingresos emitidos": {},
    "Devoluciones, descuentos y bonificaciones en egresos emitidos": {},
    "Compras y gastos no considerados en el pre llenado": {"prefix": "(+) "},
    "No considerados en el pre llenado Ingresos PUE": {},
    "No considerados en el pre llenado Pagos": {},
    "Facturas de egresos recibidas por compras y gastos": {"prefix": "(-) "},
    "Compras y gastos": {"prefix": "(=) ", "bold": True},
    "Deducciones autorizadas sin inversiones": {"prefix": "(=) ", "bold": True},
    "Adquisiciones por concepto de inversiones": {},
}

# Column header con traduccion CFDI
ISR_CFDI = {
    "ConteoCFDIs": "Conteo de CFDIs",
    "SubTotal": "Subtotal",
    "DescuentoMXN": "Descuentos",
    "NetoMXN": "Neto",
    "RetencionesISRMXN": "Retenciones ISR",
}

# Column header con traduccion DoctoRelacionado
ISR_DOCTORELACIONADO = {
    "ConteoCFDIs": "Conteo de CFDIs",
    "BaseIVA16": "Base IVA 16",
    "BaseIVA8": "Base IVA 8",
    "BaseIVA0": "Base IVA 0",
    "BaseIVAExento": "Base IVA Exento",
    "Neto": "Neto",
    "RetencionesISR": "Retenciones ISR",
}


def export_total_isr_page(isr_data: dict) -> bytes:
    """
    Genera un archivo Excel con los totales de ISR.

    Args:
        isr_data: Diccionario con los datos de ISR

    Returns:
        bytes: El workbook convertido a bytes
    """
    # Crear mapa de conceptos O(n)
    concept_map = {
        item["Concepto"]: item
        for concept in isr_data.get("totals_table", [])
        for item in [concept] + concept.get("concepts", [])
    }

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Totales"
    ws.append(["", "Conteo de CFDIs", "Importe", "ISR Retenido a Cargo"])

    # Procesar y escribir filas O(m) donde m = conceptos ordenados
    for concept_name, config in ISR_CONCEPT_CONFIG.items():
        if concept_name not in concept_map:
            continue

        item = concept_map[concept_name]

        # Formatear nombre con prefijo y/o porcentaje
        formatted_name = config.get("prefix", "") + concept_name

        suffix = config.get("suffix_dynamic")
        if suffix and suffix in item:
            formatted_name += f" {int(item[suffix] * 100)}%"

        # Agregar fila
        ws.append(
            (
                formatted_name,
                item.get("ConteoCFDIs", 0),
                item.get("Importe", 0),
                item.get("isr_cargo", 0),
            )
        )

        # Aplicar negrita si corresponde
        if config.get("bold"):
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    return _workbook_to_bytes(wb)


def _workbook_to_bytes(workbook: Workbook) -> bytes:
    """Convert openpyxl Workbook to bytes"""
    xlsx_bytes = BytesIO()
    workbook.save(xlsx_bytes)
    xlsx_bytes.seek(0)
    return xlsx_bytes.read()


def _export_isr_generic(
    company_session: Session,
    export_payload: dict,
    controller_class,
    column_mapping: dict,
    total_key: str,
    build_total_query_func,
) -> bytes:
    """
    Método genérico para exportar datos ISR a XLSX con hoja de totales.

    Args:
        company_session: Sesión de SQLAlchemy
        export_payload: Payload con configuración de exportación
        controller_class: Clase controller (CFDIController o DoctoRelacionadoController)
        column_mapping: Mapeo de nombres de columnas
        total_key: Key del payload para totales ("total_cfdi" o "total_pagos")
        build_total_query_func: Función para construir query de totales

    Returns:
        bytes: El workbook convertido a bytes
    """
    from chalicelib.blueprints.common import get_search_attrs

    search_attrs = get_search_attrs(export_payload)
    search_attrs["limit"] = None
    search_attrs["offset"] = None

    query = controller_class._get_search_query(
        domain=search_attrs["domain"],
        order_by=search_attrs["order_by"],
        limit=search_attrs["limit"],
        offset=search_attrs["offset"],
        active=search_attrs["active"],
        fuzzy_search=search_attrs["fuzzy_search"],
        fields=export_payload.get("fields", {}),
        session=company_session,
    )

    workbook = query_to_xlsx(query)
    workbook.active.title = "CFDI"

    # Hoja 2: Totales
    domain_totales = export_payload[total_key]["domain_totales"]
    fields_totales = export_payload[total_key]["fields_totales"]
    query_totales = build_total_query_func(company_session, domain_totales, fields_totales)
    result_totales_page_tupla = query_totales.first()

    ws_totales = workbook.create_sheet(title="Totales")
    headers = ["Conteo de CFDIs"] + [column_mapping[field] for field in fields_totales]
    ws_totales.append(headers)
    if result_totales_page_tupla:
        ws_totales.append(tuple(result_totales_page_tupla))

    return _workbook_to_bytes(workbook)


# METODOS AUXILIARES
def create_export_record(company_session, json_body):
    """Create export record in database"""
    period = datetime.fromisoformat(json_body["period"])

    export_request = ExportRequestORM(
        start=period.isoformat(),
        displayed_name=json_body["displayed_name"],
        domain="",
        export_data_type=ExportRequestORM.ExportDataType.ISR,
        format="XLSX",
        download_type=(DownloadType.ISSUED if json_body["issued"] else DownloadType.RECEIVED).value,
        external_request=json_body["yearly"],
        file_name=json_body["export_data"]["file_name"],
    )
    company_session.add(export_request)
    company_session.commit()
    return export_request


def save_export_to_s3(company_session, export_bytes, export_request, export_data):
    """Save export file to S3"""
    exporter = CFDIExporter(
        company_session=company_session,
        cfdi_export_repo=CFDIExportRepositorySA(session=company_session),
        bus=get_global_bus(),
    )
    exporter.export_repo_s3.save(export_bytes, export_request, export_data)

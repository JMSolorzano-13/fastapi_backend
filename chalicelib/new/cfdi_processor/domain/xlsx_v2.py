import csv
import io
import json
import tempfile
from dataclasses import dataclass
from datetime import date
from logging import ERROR

from dateutil.parser import parse as parse_date
from openpyxl import Workbook
from sqlalchemy import Boolean, DateTime, Integer, Numeric, String
from sqlalchemy.dialects.postgresql import UUID
from sqlalchemy.orm import Session

from chalicelib.controllers.cfdi import CFDIController
from chalicelib.controllers.cfdi_excluded import ExcludedCFDIController
from chalicelib.controllers.cfdi_utils.parsers import get_complementos
from chalicelib.controllers.common import CommonController
from chalicelib.logger import DEBUG, log
from chalicelib.modules import Modules
from chalicelib.new.cfdi.domain.cfdi_resume import EXERCISE, FILTERED
from chalicelib.new.config.infra.envars.control import CSV_COLUMN_LIMIT_MB
from chalicelib.new.iva import IVAGetter
from chalicelib.new.query.domain.cfdi_to_dict import CFDIDictFromXMLParser
from chalicelib.new.query.infra.copy_query import copy_query
from chalicelib.new.shared.domain.enums import Tax
from chalicelib.new.translations.resume_field_headers import headers_by_resume_type
from chalicelib.new.translations.xlsx_fields import (
    common_fields,
    egreso,
    ingreso,
    nomina,
    pagos,
    traslados,
)
from chalicelib.schema.models.tenant import CFDI as CFDIORM
from chalicelib.schema.UserDefinedType.mx_amount import MXAmount

TYPE_CONVERSIONS = {
    String: str,
    UUID: str,
    Integer: lambda x, ctx=None: int(x) if x.isdigit() else None,
    Numeric: lambda x, ctx=None: float(x)
    if x.replace(".", "", 1).replace("-", "").isdigit()
    else None,
    MXAmount: lambda x, ctx=None: float(x) if x.replace(".", "", 1).isdigit() else None,
    Boolean: lambda x, ctx=None: x.lower() in ("true", "1", "t"),
    DateTime: lambda x, ctx=None: safe_parse_datetime(x, ctx),
}


def safe_parse_datetime(x, value_context=None):
    if not x:
        return None
    try:
        return parse_date(x)
    except Exception as e:
        log(
            Modules.EXPORT,
            ERROR,
            "DATE_PARSE_ERROR",
            {
                "input_value": x,
                "expected_format": "flexible ISO8601",
                "context": value_context or {},
                "exception": str(e),
            },
        )
        raise


@dataclass
class ExportV2:
    session: Session
    # Magic number for column width
    COL_RESIZE_MAGIC_NUMBER = 1.1

    # Row range for calculating column width during auto-adjustment.
    # MIN_ROW: Starting row (headers).
    # MAX_ROW: Limits to 10 rows for balance between accuracy and performance.
    MIN_ROW = 1
    MAX_ROW = 10

    def export(self, query, body, extra_fields, resume_type, column_types_override=None):
        csv.field_size_limit(CSV_COLUMN_LIMIT_MB * 1024 * 1024)
        # MAP HEADERS TRANSLATE
        tipo_comprobante = next((d[2] for d in body["domain"] if d[0] == "TipoDeComprobante"), None)
        field_mappings = self.get_field_mappings(tipo_comprobante)

        export_concepts = False
        export_nomina_complement = False
        query_str = str(query.statement.compile(compile_kwargs={"literal_binds": True}))
        log(
            Modules.EXPORT,
            DEBUG,
            "EXPORT_V2_QUERY",
            {
                "company_identifier": body["domain"][0][2],
                "query": query_str,
                "body": body,
            },
        )
        with tempfile.NamedTemporaryFile(mode="wb", suffix=".csv") as temp_file:
            copy_query(self.session, query_str, temp_file)
            temp_file_path = temp_file.name

            wb = Workbook()
            ws = wb.active
            ws.title = "CFDI"

            with open(temp_file_path, encoding="utf-8") as temp_file_read:
                reader = csv.reader(temp_file_read)
                header = next(reader)

                new_header = [field_mappings.get(field, field) for field in header + extra_fields]

                if "Conceptos" in new_header:
                    new_header.remove("Conceptos")
                    export_concepts = True

                if "xml_content_text" in new_header:
                    new_header.remove("xml_content_text")
                    export_nomina_complement = True

                ws.append(new_header)

                if column_types_override:
                    column_types = [column_types_override.get(col, String) for col in new_header]
                else:
                    column_types = self.get_column_types(header, CFDIORM)

                for row in reader:
                    converted_row = [
                        self.convert_value(value, column_types[i], header[i], row)
                        for i, value in enumerate(row)
                    ]
                    if export_concepts:
                        conceptos_value = converted_row.pop()
                        self.process_conceptos(conceptos_value, converted_row, ws, extra_fields)
                    elif export_nomina_complement:
                        xml_content_text = converted_row.pop()
                        self.process_nomina_complement(
                            xml_content_text, converted_row, ws, extra_fields
                        )
                    else:
                        ws.append(converted_row)

            self._adjust_col_size(ws)

            self.sheet_resume_(body, wb, resume_type)
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)

        return excel_bytes.getvalue()

    def sheet_resume_(self, body, wb, resume_type):
        cfdi_resume = CFDIController.resume(
            domain=body["domain"],
            fuzzy_search=body["fuzzy_search"],
            session=self.session,
            context=None,
            resume_type=resume_type,
        )

        fields_to_use = headers_by_resume_type[resume_type]

        ws_totales = wb.create_sheet(title="Totales")
        headers = ("Tipe",) + tuple(fields_to_use.values())
        ws_totales.append(headers)
        keys = fields_to_use.keys()
        filtered = ("Periodo",) + tuple(cfdi_resume[FILTERED][k] for k in keys)
        ws_totales.append(filtered)
        exercise = ("Acumulado",) + tuple(cfdi_resume[EXERCISE][k] for k in keys)
        ws_totales.append(exercise)

        self._adjust_col_size(ws_totales)

    def get_column_types(self, header, model):
        column_types = []
        for column in header:
            base_field = getattr(model, column, None)
            if column == "balance":
                column_types.append(Numeric)
            elif base_field:
                field_type = base_field.type.__class__
                column_types.append(field_type)
            else:
                relational_field, field_related = CommonController._get_relational_fields(
                    model, [column]
                )
                field_type = field_related[0].type.__class__ if field_related else String
                column_types.append(field_type)
        return column_types

    def convert_value(self, value, field_type, column_name=None, full_row=None):
        conversion_func = TYPE_CONVERSIONS.get(field_type, lambda x: x)
        if field_type == DateTime:
            return conversion_func(value, {"column": column_name, "row": full_row})
        return conversion_func(value)

    def _adjust_col_size(self, ws, max_width=100):
        for column_cells in ws.iter_cols(min_row=self.MIN_ROW, max_row=self.MAX_ROW):
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                length * self.COL_RESIZE_MAGIC_NUMBER, max_width
            )

    def get_field_mappings(self, tipo_comprobante):
        comprobante_fields = {
            "P": {**common_fields, **pagos},
            "N": {**common_fields, **nomina},
            "E": {**common_fields, **egreso},
            "I": {**common_fields, **ingreso},
            "T": {**common_fields, **traslados},
        }

        return comprobante_fields.get(tipo_comprobante, common_fields)

    def process_conceptos(self, conceptos_value, converted_row, ws, extra_fields):
        """
        Procesa y mapea el campo 'Conceptos', manejando también el caso en que el valor sea vacío.
        Si 'Conceptos' contiene información, cada concepto se mapea y agrega a la hoja de trabajo.
        Si está vacío, se agrega directamente 'converted_row' a la hoja de trabajo.
        """
        if not conceptos_value:  # Caso cuando Conceptos está vacío o es ""
            ws.append(converted_row)
            return

        try:
            conceptos = json.loads(conceptos_value)
        except json.JSONDecodeError:
            ws.append(converted_row)
            return

        if "Concepto" in conceptos:
            conceptos["Concepto"] = (
                [conceptos["Concepto"]]
                if isinstance(conceptos["Concepto"], dict)
                else conceptos["Concepto"]
            )
            for concepto in conceptos["Concepto"]:
                mapped_data = self._data_map_for_conceptos(concepto, converted_row, extra_fields)
                ws.append(mapped_data)
        else:
            ws.append(converted_row)

    def _data_map_for_conceptos(self, concepto, converted_row, conceptos_fields):
        keys_to_convert_to_float = [
            "@Cantidad",
            "@Unidad",
            "@ValorUnitario",
            "@Importe",
            "@Descuento",
        ]
        dic_concept = self.parse_and_map_conceptos(
            concepto, conceptos_fields, keys_to_convert_to_float
        )

        # NOTA: Se hizo de esta forma porque puede llegar None Impuestos
        # Será necesario revisar en el parseador de XML una forma de evitar hacer esto
        # librería xmltodict
        impuestos = concepto.get("Impuestos") or {}
        traslados = (impuestos.get("Traslados") or {}).get("Traslado", [])
        retenciones = (impuestos.get("Retenciones") or {}).get("Retencion", [])

        traslados = [traslados] if isinstance(traslados, dict) else traslados
        retenciones = [retenciones] if isinstance(retenciones, dict) else retenciones

        dic_concept.update(self.parse_taxes(traslados, "Traslado", conceptos_fields))
        dic_concept.update(self.parse_taxes(retenciones, "Retencion", conceptos_fields))

        return converted_row + list(dic_concept.values())

    def parse_and_map_conceptos(self, concepto, fields_to_map, keys_to_float):
        mapped_data = {field: "" for field in fields_to_map}  # Inicializar mapeo
        for key in fields_to_map:
            if key == "@CuentaPredial.Numero" and "CuentaPredial" in concepto:
                cuenta_predial = concepto["CuentaPredial"]
                if isinstance(cuenta_predial, dict):
                    cuenta_predial = [cuenta_predial]
                cuenta_predial_list = []
                for cuenta in cuenta_predial:
                    if isinstance(cuenta, dict):
                        value = cuenta.get("@Numero", "")
                        if value:
                            cuenta_predial_list.append(value)

                mapped_data[key] = ", ".join(cuenta_predial_list)
                continue
            if key in concepto:
                mapped_data[key] = concepto[key]
            if key in keys_to_float:
                mapped_data[key] = try_parse_float(mapped_data[key])
        return mapped_data

    def parse_taxes(self, tax_list, tax_type, conceptos_fields):
        tax_data = {}
        plural_tax_type = {
            "Traslado": "Traslados",
            "Retencion": "Retenciones",
        }
        for tax in tax_list:
            impuesto_key = f"@Impuestos.{plural_tax_type[tax_type]}.{tax_type}@{get_tax_name(tax.get('@Impuesto', ''))}"  # noqa: E501
            for field in tax:
                mapped_key = f"{impuesto_key}{field}"
                mapped_key = "@" + mapped_key[1:].replace("@", ".")
                if mapped_key in conceptos_fields:
                    tax_data[mapped_key] = try_parse_float(tax.get(field))
        return tax_data

    def get_tax_name(tax_code):
        try:
            return Tax(tax_code).name  # Retorna el nombre (ISR, IVA, etc.)
        except ValueError:
            return "UNKNOWN"

    def process_nomina_complement(self, xml_content, converted_row, ws, extra_fields):
        if not xml_content:
            ws.append(converted_row)
            return
        fields_map = {
            "Percepciones": {"key": "Percepcion", "label": "Percepcion", "calculate_importe": True},
            "Deducciones": {"key": "Deduccion", "label": "Deduccion", "calculate_importe": False},
            "OtrosPagos": {"key": "OtroPago", "label": "Otro Pago", "calculate_importe": False},
        }
        keys_convert_to_float = ["@Importe", "@ImporteGravado", "@ImporteExento"]

        dict_xml = CFDIDictFromXMLParser().get_dict_from_xml(xml_content)

        nomina = get_complementos(dict_xml, "Nomina")

        nomina = nomina[0] if nomina else {}

        if not nomina:
            ws.append(converted_row)
            return

        for section, config in fields_map.items():
            items = nomina.get(section, {}).get(config["key"], [])
            if isinstance(items, dict):
                items = [items]

            for item in items:
                self._append_nomina_row(
                    item, converted_row, ws, extra_fields, config, keys_convert_to_float
                )

    def _append_nomina_row(
        self, item, converted_row, ws, extra_fields, config, keys_convert_to_float
    ):
        row = []

        if config["key"] == "Percepcion":
            item["@TipoComplemento"] = item.get("@TipoPercepcion", "")
        elif config["key"] == "Deduccion":
            item["@TipoComplemento"] = item.get("@TipoDeduccion", "")
        elif config["key"] == "OtroPago":
            item["@TipoComplemento"] = item.get("@TipoOtroPago", "")

        for target_key in extra_fields:
            if target_key == "@Tipo":
                row.append(config["label"])
            elif target_key == "@Importe" and config["calculate_importe"]:
                importe_gravado = try_parse_float(item.get("@ImporteGravado", "0"))
                importe_exento = try_parse_float(item.get("@ImporteExento", "0"))
                row.append(importe_gravado + importe_exento)
            else:
                value = item.get(target_key, "")
                row.append(try_parse_float(value) if target_key in keys_convert_to_float else value)
        ws.append(converted_row + row)


@dataclass
class Export_iva(ExportV2):
    def export_iva(self, query, column_types, body):
        csv.field_size_limit(CSV_COLUMN_LIMIT_MB * 1024 * 1024)

        query_str = str(query.statement.compile(compile_kwargs={"literal_binds": True}))
        with tempfile.NamedTemporaryFile(mode="wb", suffix=".csv") as temp_file:
            copy_query(self.session, query_str, temp_file)
            temp_file_path = temp_file.name

            wb = Workbook()
            ws = wb.active
            ws.title = "CFDI"

            with open(temp_file_path, encoding="utf-8") as temp_file_read:
                reader = csv.reader(temp_file_read)
                header = next(reader)

                ws.append(header)

                for row in reader:
                    converted_row = [
                        self.convert_value(value, column_types[i]) for i, value in enumerate(row)
                    ]
                    ws.append(converted_row)

            self._adjust_col_size(ws)
            if body["iva"] == "all":
                self.sheet_resume_iva(wb=wb, body=body)
            excel_bytes = io.BytesIO()
            wb.save(excel_bytes)
            excel_bytes.seek(0)

        return excel_bytes.getvalue()

    def generate_query_excluded_iva_and_all_iva(
        self,
        body,
        active: bool = True,
    ):
        order_by = body["order_by"]
        domain = body["domain"]

        query, _ = ExcludedCFDIController._get_search_query_and_count(
            session=self.session,
            domain=domain,
            fields={},
            order_by=order_by,
        )

        columns_to_keep = [
            column["expr"]
            for column in query.column_descriptions
            if column["name"]
            not in [
                "is_issued",
                "Version",
                "DR-ImpPagadoMXN",
                "FechaPago",
                "DR-RetencionesIVAMXN",
                "DR-TrasladosIVAMXN",
                "DR-IVATrasladado8",
                "DR-IVATrasladado16",
                "DR-BaseIVAExento",
                "DR-BaseIVA0",
                "DR-BaseIVA8",
                "DR-BaseIVA16",
                "DR-ExcludeFromIVA",
                "DR-Serie",
                "DR-Folio",
                "DR-Identifier",
                "DR-FormaPagoCode",
                "DR-FormaPagoName",
                "ExcludeFromIVA",
                "NumParcialidad",
            ]
        ]
        query = query.with_entities(*columns_to_keep)

        order_and_names = {
            "Fecha": "Fecha de emisión",
            "PaymentDate": "Fecha de pago",
            "UUID": "UUID",
            "Serie": "Serie",
            "Folio": "Folio",
            "RfcEmisor": "RFC emisor",
            "NombreEmisor": "Emisor",
            "TipoDeComprobante": "Tipo de comprobante",
            "UsoCFDIReceptor": "Uso de CFDI",
            "FormaPago": "Forma de pago código",
            "FormaPagoName": "Forma de pago",
            "MetodoPago": "Método de pago",
            "BaseIVA16": "Base IVA 16%",
            "BaseIVA8": "Base IVA 8%",
            "BaseIVA0": "Base IVA 0%",
            "BaseIVAExento": "Base IVA exento",
            "IVATrasladado16": "IVA 16%",
            "IVATrasladado8": "IVA 8%",
            "TrasladosIVA": "IVA acreditable total",
            "RetencionesIVA": "Retenciones IVA",
            "Total": "Total",
            "DR-UUID": "UUID pago",
        }

        columns_with_labels = []
        column_types = []

        for key in order_and_names:
            for column in query.column_descriptions:
                if column["name"] == key:
                    column_class = DateTime if key == "PaymentDate" else type(column["type"])
                    column_types.append(column_class)
                    columns_with_labels.append(column["expr"].label(order_and_names[key]))

        query = query.with_entities(*columns_with_labels)

        return query, column_types

    def sheet_resume_iva(self, wb, body):
        period = date.fromisoformat(body["period"])

        getter = IVAGetter(self.session)

        iva = getter.get_iva(period)

        ws_totales = wb.create_sheet(title="Totales")

        headers = [
            "",
            "Conteo de CFDIs",
            "Base IVA 16%",
            "Base IVA 8%",
            "Base IVA 0%",
            "Base IVA exento",
            "IVA acreditable 16%",
            "IVA acreditable 8%",
            "IVA acreditable total",
            "Retenciones IVA",
        ]

        ws_totales.append(headers)

        facturas_contado = [
            "Facturas de contado",
            iva["period"]["creditable"]["i_tra"]["qty"],
            iva["period"]["creditable"]["i_tra"]["BaseIVA16"],
            iva["period"]["creditable"]["i_tra"]["BaseIVA8"],
            iva["period"]["creditable"]["i_tra"]["BaseIVA0"],
            iva["period"]["creditable"]["i_tra"]["BaseIVAExento"],
            iva["period"]["creditable"]["i_tra"]["IVATrasladado16"],
            iva["period"]["creditable"]["i_tra"]["IVATrasladado8"],
            iva["period"]["creditable"]["i_tra"]["TrasladosIVAMXN"],
            iva["period"]["creditable"]["i_tra"]["RetencionesIVAMXN"],
        ]
        facturas_credito = [
            "Facturas de crédito",
            iva["period"]["creditable"]["p_tra"]["qty"],
            iva["period"]["creditable"]["p_tra"]["BaseIVA16"],
            iva["period"]["creditable"]["p_tra"]["BaseIVA8"],
            iva["period"]["creditable"]["p_tra"]["BaseIVA0"],
            iva["period"]["creditable"]["p_tra"]["BaseIVAExento"],
            iva["period"]["creditable"]["p_tra"]["IVATrasladado16"],
            iva["period"]["creditable"]["p_tra"]["IVATrasladado8"],
            iva["period"]["creditable"]["p_tra"]["total"],
            iva["period"]["creditable"]["p_tra"]["RetencionesIVAMXN"],
        ]
        totales = [
            "Totales",
            iva["period"]["creditable"]["p_tra"]["qty"]
            + iva["period"]["creditable"]["i_tra"]["qty"],
            iva["period"]["creditable"]["p_tra"]["BaseIVA16"]
            + iva["period"]["creditable"]["i_tra"]["BaseIVA16"],
            iva["period"]["creditable"]["p_tra"]["BaseIVA8"]
            + iva["period"]["creditable"]["i_tra"]["BaseIVA8"],
            iva["period"]["creditable"]["p_tra"]["BaseIVA0"]
            + iva["period"]["creditable"]["i_tra"]["BaseIVA0"],
            iva["period"]["creditable"]["p_tra"]["BaseIVAExento"]
            + iva["period"]["creditable"]["i_tra"]["BaseIVAExento"],
            iva["period"]["creditable"]["p_tra"]["IVATrasladado16"]
            + iva["period"]["creditable"]["i_tra"]["IVATrasladado16"],
            iva["period"]["creditable"]["p_tra"]["IVATrasladado8"]
            + iva["period"]["creditable"]["i_tra"]["IVATrasladado8"],
            iva["period"]["creditable"]["p_tra"]["total"]
            + iva["period"]["creditable"]["i_tra"]["total"],
            iva["period"]["creditable"]["p_tra"]["RetencionesIVAMXN"]
            + iva["period"]["creditable"]["i_tra"]["RetencionesIVAMXN"],
        ]
        notas_credito = [
            "Notas de crédito",
            iva["period"]["creditable"]["credit_notes"]["qty"],
            iva["period"]["creditable"]["credit_notes"]["BaseIVA16"],
            iva["period"]["creditable"]["credit_notes"]["BaseIVA8"],
            iva["period"]["creditable"]["credit_notes"]["BaseIVA0"],
            iva["period"]["creditable"]["credit_notes"]["BaseIVAExento"],
            iva["period"]["creditable"]["credit_notes"]["IVATrasladado16"],
            iva["period"]["creditable"]["credit_notes"]["IVATrasladado8"],
            iva["period"]["creditable"]["credit_notes"]["TrasladosIVAMXN"],
            iva["period"]["creditable"]["credit_notes"]["RetencionesIVAMXN"],
        ]

        ws_totales.append(facturas_contado)
        ws_totales.append(facturas_credito)
        ws_totales.append(totales)
        ws_totales.append(notas_credito)

        self._adjust_col_size(ws_totales)


def try_parse_float(value):
    try:
        return float(value)
    except ValueError:
        return value


def get_tax_name(tax_code):
    try:
        return Tax(tax_code).name  # Retorna el nombre (ISR, IVA, etc.)
    except ValueError:
        return "UNKNOWN"

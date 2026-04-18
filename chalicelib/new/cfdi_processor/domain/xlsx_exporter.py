import contextlib
import json
from collections.abc import Iterable
from datetime import date, datetime
from io import BytesIO
from typing import Any

from lxml import etree
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from chalicelib.logger import EXCEPTION, WARNING, log
from chalicelib.modules import Modules
from chalicelib.new.cfdi.domain.cfdi_resume import EXERCISE, FILTERED, CFDIResume
from chalicelib.new.model_serializer.app.model_serializer import ModelSerializer
from chalicelib.new.shared.domain.enums import Tax, TaxFactor
from chalicelib.new.translations.resume_field_headers import headers_by_resume_type
from chalicelib.new.translations.xlsx_fields import XLSXFields
from chalicelib.schema.models.tenant import CFDI as CFDIORM

PrimitiveType = str | int | float | bool | date | datetime

dotted_fields_exceptions = ["paid_by", "cfdi_related"]


class XMLParser:
    def __init__(self, docto_pago):
        if not docto_pago.cfdi_origin or docto_pago.cfdi_origin.xml_content is None:
            self.xml_content = b""
        else:
            self.xml_content = docto_pago.cfdi_origin.xml_content.encode()

        self.docto_pago = docto_pago
        self.namespaces_pago20 = {
            "cfdi": "http://www.sat.gob.mx/cfd/4",
            "pago20": "http://www.sat.gob.mx/Pagos20",
        }
        self.namespaces_pago10 = {
            "cfdi": "http://www.sat.gob.mx/cfd/3",
            "pago10": "http://www.sat.gob.mx/Pagos",
        }

    def namespaces(self):
        return {**self.namespaces_pago10, **self.namespaces_pago20}

    def parse_xml(self):
        if not self.xml_content or not self.xml_content.strip():
            return None
        try:
            root = etree.fromstring(self.xml_content)
            return root
        except etree.XMLSyntaxError as e:
            log(
                Modules.PROCESS_XML,
                EXCEPTION,
                "XML_SYNTAX_ERROR",
                {"xml": self.xml_content, "exception": e},
            )
            return None

    def search_id_document(self, docto_pago):
        id_document = docto_pago.UUID_related.strip()
        num_parcialidad = docto_pago.NumParcialidad
        normalized_id_document = id_document.upper()
        root = self.parse_xml()
        if root is None:
            return None

        docto_relacionados = root.findall(
            ".//pago10:DoctoRelacionado", self.namespaces()
        ) + root.findall(".//pago20:DoctoRelacionado", self.namespaces())

        docto_relacionado = next(
            (
                node
                for node in docto_relacionados
                if node.get("IdDocumento", "").upper() == normalized_id_document
                and node.get("NumParcialidad") == str(num_parcialidad)
            ),
            None,
        )
        return docto_relacionado


class PropertiesNodeParentBase(BaseModel):
    FechaPago: datetime = datetime.now()
    FormaDePagoP: str = ""
    MonedaP: str = ""
    Monto: float = 0.0
    TipoCambioP: str = ""


class PropertiesNodeIdDocumentBase(BaseModel):
    EquivalenciaDR: float = 1.0
    Folio: str = ""
    ImpPagado: float = 0.0
    ImpPagadoMXN: float = 0.0
    ImpSaldoAnt: float = 0.0
    ImpSaldoInsoluto: float = 0.0
    MonedaDR: str = ""
    NumParcialidad: int = 0
    ObjetoImpDR: str = ""
    Serie: str = ""


class PropertiesTaxData(BaseModel):
    BaseIva16MXN: float = 0.0
    Iva16MXN: float = 0.0
    BaseIva08MXN: float = 0.0
    Iva08MXN: float = 0.0
    BaseIva0MXN: float = 0.0
    Iva0MXN: float = 0.0
    BaseIvaExentoMXN: float = 0.0
    RetencionesIsrMXN: float = 0.0
    RetencionesIvaMXN: float = 0.0
    BaseIEPS: float = 0.0
    Ieps: float = 0.0
    TipoFactorIEPS: str = ""
    TasaCuotaIEPS: float = 0.0


class CompositeProperties(BaseModel):
    node_parent: PropertiesNodeParentBase
    node_document: PropertiesNodeIdDocumentBase
    tax_data: PropertiesTaxData


class PropertiesNodeParent:
    def __init__(self, docto_relacionado):
        self.docto_relacionado = docto_relacionado
        self.properties = PropertiesNodeParentBase()

    def get_attrs(self):
        if not self.docto_relacionado:
            self._set_default_properties()
            log(
                Modules.PROCESS_XML,
                WARNING,
                "DOCTO_RELACIONADO_NONE",
                {
                    "docto_relacionado": self.docto_relacionado,
                },
            )
            return self.properties

        self.properties.FechaPago = self.docto_relacionado.getparent().get("FechaPago", "")
        self.properties.FormaDePagoP = self.docto_relacionado.getparent().get("FormaDePagoP", "")
        self.properties.MonedaP = self.docto_relacionado.getparent().get("MonedaP", "")
        self.properties.Monto = float(self.docto_relacionado.getparent().get("Monto", 0))
        self.properties.TipoCambioP = float(
            self.docto_relacionado.getparent().get("TipoCambioP", 1)
        )
        return self.properties

    def _set_default_properties(self):
        """Helper method to set default property values"""
        self.properties.FechaPago = ""
        self.properties.FormaDePagoP = ""
        self.properties.MonedaP = ""
        self.properties.Monto = 0
        self.properties.TipoCambioP = 0


class PropertiesNodeIdDocument:
    def __init__(self, docto_relacionado):
        self.docto_relacionado = docto_relacionado
        self.properties = PropertiesNodeIdDocumentBase()

    def get_attrs(self):
        if not self.docto_relacionado:
            self._set_default_properties()
            log(
                Modules.PROCESS_XML,
                WARNING,
                "DOCTO_RELACIONADO_NONE_2",
                {
                    "docto_relacionado": self.docto_relacionado,
                },
            )
            return self.properties

        self.properties.EquivalenciaDR = float(self.docto_relacionado.get("EquivalenciaDR", 1))
        self.properties.Folio = self.docto_relacionado.get("Folio", "")
        self.properties.ImpPagado = float(self.docto_relacionado.get("ImpPagado", 0))
        self.properties.ImpSaldoAnt = float(self.docto_relacionado.get("ImpSaldoAnt", 0))
        self.properties.ImpSaldoInsoluto = float(self.docto_relacionado.get("ImpSaldoInsoluto", 0))
        self.properties.MonedaDR = self.docto_relacionado.get("MonedaDR", "")
        self.properties.NumParcialidad = int(self.docto_relacionado.get("NumParcialidad", 0))
        self.properties.ObjetoImpDR = self.docto_relacionado.get("ObjetoImpDR", "")
        self.properties.Serie = self.docto_relacionado.get("Serie", "")
        return self.properties

    def _set_default_properties(self):
        """Helper method to set default property values"""
        self.properties.EquivalenciaDR = 1
        self.properties.Folio = ""
        self.properties.ImpPagado = 0
        self.properties.ImpSaldoAnt = 0
        self.properties.ImpSaldoInsoluto = 0
        self.properties.MonedaDR = ""
        self.properties.NumParcialidad = 0
        self.properties.ObjetoImpDR = ""
        self.properties.Serie = ""


class TrasladosManager:
    def __init__(self, docto_relacionado, namespaces):
        self.docto_relacionado = docto_relacionado
        self.namespaces = namespaces

    def get_data(self, docto_relacionado, namespaces) -> list:
        if not docto_relacionado:
            log(
                Modules.PROCESS_XML,
                WARNING,
                "DOCTO_RELACIONADO_NONE_3",
                {
                    "docto_relacionado": self.docto_relacionado,
                },
            )
            return [
                {"BaseDR": 0, "ImporteDR": 0, "ImpuestoDR": 0, "TasaOCuotaDR": 0, "TipoFactorDR": 0}
            ]

        list_tralados = []
        traslados_dr = docto_relacionado.findall(
            ".//pago20:TrasladosDR/pago20:TrasladoDR", namespaces
        )

        for traslado in traslados_dr:
            new_node_traslado = {}
            new_node_traslado["BaseDR"] = float(traslado.get("BaseDR", 0))
            new_node_traslado["ImporteDR"] = float(traslado.get("ImporteDR", 0))
            new_node_traslado["ImpuestoDR"] = traslado.get("ImpuestoDR", 0)
            new_node_traslado["TasaOCuotaDR"] = float(traslado.get("TasaOCuotaDR", 0))
            new_node_traslado["TipoFactorDR"] = traslado.get("TipoFactorDR", 0)
            list_tralados.append(new_node_traslado)
        return list_tralados


class RetencionesManager:
    def __init__(self, docto_relacionado, namespaces):
        self.docto_relacionado = docto_relacionado
        self.namespaces = namespaces

    def get_data(self, docto_relacionado, namespaces) -> list:
        if not docto_relacionado:
            log(
                Modules.PROCESS_XML,
                WARNING,
                "DOCTO_RELACIONADO_NONE_4",
                {
                    "docto_relacionado": self.docto_relacionado,
                },
            )
            return [
                {"BaseDR": 0, "ImporteDR": 0, "ImpuestoDR": 0, "TasaOCuotaDR": 0, "TipoFactorDR": 0}
            ]
        list_retenciones = []
        retenciones_dr = docto_relacionado.findall(
            ".//pago20:RetencionesDR/pago20:RetencionDR", namespaces
        )
        for retencion in retenciones_dr:
            new_node_retencion = {}
            new_node_retencion["BaseDR"] = float(retencion.get("BaseDR", 0))
            new_node_retencion["ImporteDR"] = float(retencion.get("ImporteDR", 0))
            new_node_retencion["ImpuestoDR"] = retencion.get("ImpuestoDR", 0)
            new_node_retencion["TasaOCuotaDR"] = float(retencion.get("TasaOCuotaDR", 0))
            new_node_retencion["TipoFactorDR"] = retencion.get("TipoFactorDR", 0)
            list_retenciones.append(new_node_retencion)
        return list_retenciones


class TaxCalculator:
    def __init__(self, properties):
        self.properties = properties
        self.equivalencia_dr = self.properties.EquivalenciaDR
        self.tipo_cambio_p = self.properties.payment_related.TipoCambioP or 1.0
        self.find_base_iva_16 = False
        self.find_base_iva_8 = False
        self.find_base_iva_0 = False

    def calculate_fields(self, data, data_traslados, data_retenciones):
        if data_traslados:
            for traslado in data_traslados:
                tax_mapping = {
                    0.16: ("BaseIva16MXN", "Iva16MXN", "find_base_iva_16"),
                    0.08: ("BaseIva08MXN", "Iva08MXN", "find_base_iva_8"),
                    0.0: ("BaseIva0MXN", "Iva0MXN", "find_base_iva_0"),
                }

                if (
                    traslado["@ImpuestoDR"] == Tax.IVA
                    and traslado["@TipoFactorDR"] == TaxFactor.TASA
                    and traslado["@TasaOCuotaDR"] in tax_mapping
                ):
                    base_attr, iva_attr, find_attr = tax_mapping[traslado["@TasaOCuotaDR"]]

                    if not getattr(self, find_attr):
                        data[base_attr] = traslado["@BaseDR"]
                        data[iva_attr] = traslado["@ImporteDR"]
                        setattr(self, find_attr, True)

                if traslado["@ImpuestoDR"] == Tax.IEPS:
                    data["BaseIEPS"] += traslado["@BaseDR"]
                    data["TipoFactorIEPS"] = traslado["@TipoFactorDR"]
                    data["TasaCuotaIEPS"] = traslado.get("@TasaOCuotaDR", 0)
                    data["Ieps"] += traslado.get("@ImporteDR", 0)
                if (
                    traslado["@ImpuestoDR"] == Tax.IVA
                    and traslado["@TipoFactorDR"] == TaxFactor.EXENTO
                ):
                    data["BaseIvaExentoMXN"] = float(traslado["@BaseDR"])
        if data_retenciones:
            for retencion in data_retenciones:
                if retencion["@ImpuestoDR"] == Tax.ISR:
                    data["RetencionesIsrMXN"] = float(retencion["@ImporteDR"])
                if retencion["@ImpuestoDR"] == Tax.IVA:
                    data["RetencionesIvaMXN"] = float(retencion["@ImporteDR"])


class DoctoRelatedProcessor:
    def __init__(self, docto_pago):
        self.docto_pago = docto_pago
        self.composite_properties = CompositeProperties(
            node_parent=PropertiesNodeParentBase(),
            node_document=PropertiesNodeIdDocumentBase(),
            tax_data=PropertiesTaxData(),
        )
        self.xml_parser = XMLParser(self.docto_pago)
        self.traslados_manager = TrasladosManager(self.docto_pago, self.xml_parser.namespaces())
        self.retenciones_manager = RetencionesManager(self.docto_pago, self.xml_parser.namespaces())

    def process(self):
        docto_relacionado = self.docto_pago
        tax_calculator = TaxCalculator(properties=docto_relacionado)
        data_traslados = docto_relacionado.TrasladosDR
        data_retenciones = docto_relacionado.RetencionesDR

        data = {
            "FechaPago": docto_relacionado.FechaPago,
            "FormaPago": docto_relacionado.payment_related.FormaDePagoP,
            "MonedaP": docto_relacionado.payment_related.MonedaP,
            "Serie": docto_relacionado.Serie,
            "Folio": docto_relacionado.Folio,
            "ObjetoImpDR": docto_relacionado.ObjetoImpDR,
            "ImpPagado": docto_relacionado.ImpPagado,
            "importe_pagado_mxn": docto_relacionado.ImpPagadoMXN,
            "EquivalenciaDR": docto_relacionado.EquivalenciaDR,
            "BaseIva16MXN": 0,
            "Iva16MXN": 0,
            "BaseIva08MXN": 0,
            "Iva08MXN": 0,
            "BaseIva0MXN": 0,
            "Iva0MXN": 0,
            "BaseIvaExentoMXN": 0,
            "RetencionesIsrMXN": 0,
            "RetencionesIvaMXN": 0,
            "BaseIEPS": 0,
            "TipoFactorIEPS": 0,
            "TasaCuotaIEPS": 0,
            "Ieps": 0,
        }

        tax_calculator.calculate_fields(data, data_traslados, data_retenciones)

        return data

    def set_values_properties(self, docto_relacionado):
        properties_node_parent = PropertiesNodeParent(docto_relacionado)
        properties_node_document = PropertiesNodeIdDocument(docto_relacionado)
        self.composite_properties.node_parent = properties_node_parent.get_attrs()
        self.composite_properties.node_document = properties_node_document.get_attrs()


class XLSXExporter:
    COL_RESIZE_MAGIC_NUMBER = 1.1
    resume_header_type = "Tipo"

    def _translate_fields(self, fields: Iterable[str]) -> Iterable[str]:
        has_duplicates = all(value in fields for value in ["FechaFiltro", "PaymentDate"])
        fieldForXLSX = XLSXFields.copy()
        if has_duplicates:
            fieldForXLSX["FechaFiltro"] = "Fecha de emisión"
        return tuple(fieldForXLSX.get(field, field) for field in fields)

    def _flatten_dict(self, d, parent_key="", sep="."):
        items = {}
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.update(self._flatten_dict(v, new_key, sep=sep))
            else:
                items[new_key] = v
        return items

    def _match_headers_with_data(self, fields, data, export_data=None) -> dict:
        new_data = {}

        # Flatten the data dict for easier access to nested fields
        flatten_data = self._flatten_dict(data)

        extra_fields = [
            "Fecha de Pago",
            "Forma de Pago",
            "FormaPagoName",
            "Moneda de pago",
            "DR - Equivalencia",
            "DR - Fecha de emisión",
            "DR - Serie",
            "DR - Folio",
            "DR - Objeto de impuesto",
            "DR - Importe pagado MXN",
            "DR - Base IVA 16%",
            "DR - Base IVA 8%",
            "DR - Base IVA 0%",
            "DR - Base IVA Exento",
            "DR - IVA 16%",
            "DR - IVA 8%",
            "DR - IVA Total",
            "DR - Base IEPS",
            "DR - Factor IEPS",
            "DR - Tasa o cuota IEPS",
            "DR - IEPS",
            "DR - Retenciones ISR",
            "DR - Retenciones IVA",
        ]
        all_fields = list(fields) + extra_fields if export_data["type"] == "doctos" else fields

        for field in all_fields:
            splitted = field.split(".")
            if export_data["type"] == "doctos":
                current_field = (
                    field if splitted[0] not in dotted_fields_exceptions else splitted[0]
                )
            else:
                current_field = (
                    splitted[-1] if splitted[0] not in dotted_fields_exceptions else splitted[0]
                )
            if field.startswith("Conceptos."):
                components = field.split(".")
                subfields = components[1:]
                special_field = "@" + "@".join(subfields[-3:])
                new_data[special_field] = data.get(field, "")
            elif field.startswith("N.Complemento."):
                new_data[field] = data.get(field, "")
            else:
                if current_field == "ExcludeFromIVA":
                    new_data[field] = "True" if data.get(current_field) else "False"
                else:
                    new_data[field] = flatten_data.get(field, data.get(current_field, ""))
        return new_data

    def _data_to_worksheet(
        self,
        worksheet: Worksheet,
        cfdis: Iterable[CFDIORM],
        fields: Iterable[str],
        export_data: dict,
        title: str = "CFDI",
    ) -> None:
        worksheet.title = title
        fields_translated = self._translate_fields(fields)

        if export_data["type"] == "doctos":

            def get_additional_fields():
                additional_fields = (
                    "Fecha de Pago",
                    "Forma de Pago",
                    "Moneda de pago",
                    "DR - Equivalencia",
                    "DR - Serie",
                    "DR - Folio",
                    "DR - Objeto de impuesto",
                    "DR - Importe pagado MXN",
                    "DR - Base IVA 16%",
                    "DR - Base IVA 8%",
                    "DR - Base IVA 0%",
                    "DR - Base IVA Exento",
                    "DR - IVA 16%",
                    "DR - IVA 8%",
                    "DR - IVA Total",
                    "DR - Base IEPS",
                    "DR - Factor IEPS",
                    "DR - Tasa o cuota IEPS",
                    "DR - IEPS",
                    "DR - Retenciones ISR",
                    "DR - Retenciones IVA",
                )
                return additional_fields

            fields_translated = get_additional_fields() + fields_translated
            custom_order = [
                "Fecha de Pago",
                "Fecha expedición",
                "UUID",
                "Serie",
                "Folio",
                "RFC emisor",
                "RFC receptor",
                "Emisor",
                "Receptor",
                "Forma de Pago",
                "Moneda de pago",
                "DR - Fecha de emisión",
                "DR - Serie",
                "DR - Folio",
                "DR - UUID",
                "Uso de CFDI",
                "DR - Objeto de impuesto",
                "DR - Moneda",
                "DR - Equivalencia",
                "DR - Numero de parcialidad",
                "DR - Importe pagado",
                "DR - Importe pagado MXN",
                "DR - Base IVA 16%",
                "DR - Base IVA 8%",
                "DR - Base IVA 0%",
                "DR - Base IVA Exento",
                "DR - IVA 16%",
                "DR - IVA 8%",
                "DR - IVA Total",
                "DR - Base IEPS",
                "DR - Factor IEPS",
                "DR - Tasa o cuota IEPS",
                "DR - IEPS",
                "DR - Retenciones ISR",
                "DR - Retenciones IVA",
            ]
            fields_translated = sorted(fields_translated, key=lambda x: custom_order.index(x))
            fields_translated = [
                "Fecha de emisión"
                if x == "Fecha expedición"
                else "DR - Uso de CFDI"
                if x == "Uso de CFDI"
                else x
                for x in fields_translated
            ]

        conceptos_fields = [
            field.replace("Conceptos.", "@") for field in fields if field.startswith("Conceptos.")
        ]
        worksheet.append(fields_translated)
        n_fields = [
            campo.replace("N.Complemento.", "@")
            for campo in fields
            if campo.startswith("N.Complemento.")
        ]
        fields = [campo for campo in fields if not campo.startswith("N.Complemento")]
        serializer = ModelSerializer(process_iterable=process_iterable)
        for record in cfdis:
            data = serializer.serialize(record, fields)

            final_data = self._match_headers_with_data(fields, data, export_data=export_data)
            if export_data["type"] == "doctos":
                pays_data = self._process_pays_data(record, fields)
                self.init_export_doctos(worksheet, fields, export_data, record, data, pays_data)
            elif export_data["type"] == "iva-doctos" or export_data["type"] == "export-iva":
                worksheet.append(tuple(final_data.values()))
            else:
                if hasattr(record, "Conceptos") and record.Conceptos and conceptos_fields:
                    conceptos = json.loads(str(record.Conceptos))
                    if isinstance(conceptos["Concepto"], dict):
                        conceptos["Concepto"] = [conceptos["Concepto"]]
                    for concepto in conceptos["Concepto"]:
                        worksheet.append(
                            self._data_map_for_conceptos(concepto, final_data, conceptos_fields)
                        )
                elif hasattr(record, "nomina") and record.nomina and n_fields:
                    if record.nomina.Percepciones:
                        percepciones = record.nomina.Percepciones[0]["Percepcion"]
                        for percepcion in percepciones:
                            worksheet.append(
                                self._data_map_for_nomina_complemento(
                                    percepcion, final_data, n_fields
                                )
                            )

                    if record.nomina.Deducciones:
                        deducciones = record.nomina.Deducciones[0]["Deduccion"]
                        for deduccion in deducciones:
                            worksheet.append(
                                self._data_map_for_nomina_complemento(
                                    deduccion, final_data, n_fields
                                )
                            )

                    if record.nomina.OtrosPagos:
                        otros_pagos = record.nomina.OtrosPagos[0]["OtroPago"]
                        for otro_pago in otros_pagos:
                            worksheet.append(
                                self._data_map_for_nomina_complemento(
                                    otro_pago, final_data, n_fields
                                )
                            )
                else:
                    worksheet.append(tuple(final_data.values()))

    def _process_pays_data(self, record, fields):
        pays_data = []
        for field in fields:
            if field.startswith("pays."):
                related_field = field.split(".")[1]
                related_values = [getattr(doc, related_field, "") for doc in record.pays]
                pays_data.append((field, related_values))
        return pays_data

    def init_export_doctos(self, worksheet, fields, export_data, record, data, pays_data):
        if not pays_data:
            final_data = self._match_headers_with_data(fields, data, export_data=export_data)
            worksheet.append(final_data)
            return

        row_data = data.copy()
        for docto in record.pays:
            processor = DoctoRelatedProcessor(docto_pago=docto)
            properties = processor.process()

            if docto.cfdi_related is not None:
                row_data["UsoCFDIReceptor"] = docto.cfdi_related.UsoCFDIReceptor
                row_data["DR - Fecha de emisión"] = docto.cfdi_related.Fecha
            else:
                row_data["UsoCFDIReceptor"] = ""
                row_data["DR - Fecha de emisión"] = ""

            row_data["pays.UUID_related"] = docto.UUID_related
            row_data["pays.NumParcialidad"] = docto.NumParcialidad
            row_data["pays.ImpPagado"] = docto.ImpPagado
            row_data["pays.MonedaDR"] = docto.MonedaDR
            row_data["Fecha de Pago"] = str(properties["FechaPago"])
            row_data["Forma de Pago"] = properties["FormaPago"]
            row_data["FormaPagoName"] = docto.cfdi_origin.c_forma_pago
            row_data["Moneda de pago"] = properties["MonedaP"]
            row_data["DR - Equivalencia"] = properties["EquivalenciaDR"]
            row_data["DR - Serie"] = properties["Serie"]
            row_data["DR - Folio"] = properties["Folio"]
            row_data["DR - Objeto de impuesto"] = properties["ObjetoImpDR"]
            row_data["DR - Importe pagado MXN"] = properties["importe_pagado_mxn"]
            row_data["DR - Base IVA 16%"] = properties.get("BaseIva16MXN")
            row_data["DR - Base IVA 8%"] = properties.get("BaseIva08MXN")
            row_data["DR - Base IVA 0%"] = properties.get("BaseIva0MXN")
            row_data["DR - Base IVA Exento"] = properties.get("BaseIvaExentoMXN")
            row_data["DR - IVA 16%"] = properties.get("Iva16MXN")
            row_data["DR - IVA 8%"] = properties.get("Iva08MXN")
            row_data["DR - IVA Total"] = properties.get("Iva16MXN") + properties.get("Iva08MXN")
            row_data["DR - Base IEPS"] = properties.get("BaseIEPS")
            row_data["DR - Factor IEPS"] = properties.get("TipoFactorIEPS")
            row_data["DR - Tasa o cuota IEPS"] = properties.get("TasaCuotaIEPS")
            row_data["DR - IEPS"] = properties.get("Ieps")
            row_data["DR - Retenciones ISR"] = properties.get("RetencionesIsrMXN")
            row_data["DR - Retenciones IVA"] = properties.get("RetencionesIvaMXN")

            ordered_row_data = [
                "Fecha de Pago",
                "Fecha",
                "UUID",
                "Serie",
                "Folio",
                "RfcEmisor",
                "RfcReceptor",
                "NombreEmisor",
                "NombreReceptor",
                "Forma de Pago",
                "Moneda de pago",
                "DR - Fecha de emisión",
                "DR - Serie",
                "DR - Folio",
                "pays.UUID_related",
                "UsoCFDIReceptor",
                "DR - Objeto de impuesto",
                "pays.MonedaDR",
                "DR - Equivalencia",
                "pays.NumParcialidad",
                "pays.ImpPagado",
                "DR - Importe pagado MXN",
                "DR - Base IVA 16%",
                "DR - Base IVA 8%",
                "DR - Base IVA 0%",
                "DR - Base IVA Exento",
                "DR - IVA 16%",
                "DR - IVA 8%",
                "DR - IVA Total",
                "DR - Base IEPS",
                "DR - Factor IEPS",
                "DR - Tasa o cuota IEPS",
                "DR - IEPS",
                "DR - Retenciones ISR",
                "DR - Retenciones IVA",
            ]
            final_data = self._match_headers_with_data(fields, row_data, export_data=export_data)
            if final_data.get("RfcEmisor", None) is None:
                ordered_row_data.remove("RfcEmisor")
                ordered_row_data.remove("NombreEmisor")
            if final_data.get("RfcReceptor", None) is None:
                ordered_row_data.remove("RfcReceptor")
                ordered_row_data.remove("NombreReceptor")
            final_data = tuple(final_data[key] for key in ordered_row_data)

            worksheet.append(final_data)

    def _data_map_for_conceptos(self, concepto, final_data, conceptos_fields):
        keys_concepto = {
            "@ClaveProdServ",
            "@NoIdentificacion",
            "@Cantidad",  #
            "@ClaveUnidad",
            "@Unidad",  #
            "@Descripcion",
            "@ValorUnitario",  #
            "@Importe",  #
            "@Descuento",  #
            "@ObjetoImp",
        }
        keys_to_convert_to_float = [
            "@Cantidad",
            "@Unidad",
            "@ValorUnitario",
            "@Importe",
            "@Descuento",
        ]
        data_concepto_arroba = {
            key: value for key, value in final_data.items() if key in keys_concepto
        }
        for key in data_concepto_arroba:
            if key in concepto:
                data_concepto_arroba[key] = concepto[key]
        data_concepto_arroba = is_str_convert_to_float(
            old_dict=data_concepto_arroba, keys_to_convert=keys_to_convert_to_float
        )
        # Map the concept fields that are not in keys_concepto
        final_data.update(data_concepto_arroba)

        # Here the tax fields are added
        impuestos = concepto.get("Impuestos") or {}
        traslados = (impuestos.get("Traslados") or {}).get("Traslado", [])
        retenciones = (impuestos.get("Retenciones") or {}).get("Retencion", [])

        traslados = (
            [traslados]
            if isinstance(traslados, dict)
            else traslados
            if isinstance(traslados, list)
            else []
        )

        retenciones = (
            [retenciones]
            if isinstance(retenciones, dict)
            else retenciones
            if isinstance(retenciones, list)
            else []
        )

        # To define the common keys
        keys_traslado_retencion = {
            "@Base",
            "@Importe",
            "@Impuesto",
            "@TasaOCuota",
            "@TipoFactor",
        }

        # Build the key mapping for each tax using the common keys
        keys_mapping = {
            Tax.ISR: keys_traslado_retencion.copy(),
            Tax.IVA: keys_traslado_retencion.copy(),
            Tax.IEPS: keys_traslado_retencion.copy(),
        }

        for traslado in traslados:
            impuesto = traslado.get("@Impuesto", "")
            impuesto_key = "Traslado@IVA" if impuesto == Tax.IVA else "Traslado@IEPS"
            if impuesto in keys_mapping:
                # Build a dictionary with the keys of the tax fields
                traslado_modificado = {
                    f"@{impuesto_key}{clave}": traslado.get(clave)
                    for clave in keys_mapping[impuesto]
                }
                # Build a dictionary with keys of concept fields that are not in keys_concepto
                nuevo_diccionario = {key for key in conceptos_fields if key not in keys_concepto}

                new_dict_formatted = {}

                for key in nuevo_diccionario:
                    partes = key.split(".")
                    if "Traslado" in partes:
                        nuevo_key = "@" + "@".join(partes[2:])
                        new_dict_formatted[nuevo_key] = final_data[nuevo_key]

                for key in new_dict_formatted:
                    if key in traslado_modificado:
                        new_dict_formatted[key] = traslado_modificado[key]
                new_dict_formatted = is_str_convert_to_float(new_dict_formatted)
                final_data.update(new_dict_formatted)

        for retencion in retenciones:
            impuesto = retencion.get("@Impuesto", "")
            impuesto_key = "Retencion@ISR" if impuesto == Tax.ISR else "Retencion@IVA"
            if impuesto in keys_mapping:
                retencion_modificado = {
                    f"@{impuesto_key}{clave}": retencion.get(clave)
                    for clave in keys_mapping[impuesto]
                }
                nuevo_diccionario = {key for key in conceptos_fields if key not in keys_concepto}

                new_dict_formatted = {}
                for key in nuevo_diccionario:
                    partes = key.split(".")
                    if "Retencion" in partes:
                        nuevo_key = "@" + "@".join(partes[2:])
                        new_dict_formatted[nuevo_key] = final_data[nuevo_key]

                for key in new_dict_formatted:
                    if key in retencion_modificado:
                        new_dict_formatted[key] = retencion_modificado[key]
                new_dict_formatted = is_str_convert_to_float(new_dict_formatted)
                final_data.update(new_dict_formatted)
        return tuple(final_data.values())

    def _data_map_for_nomina_complemento(self, complemento, final_data, n_fields):
        commons_keys_dict = {"@ImporteGravado": "", "@ImporteExento": ""}
        keys_to_convert_to_float = [
            "@Importe",
            "@ImporteGravado",
            "@ImporteExento",
        ]
        if "@TipoPercepcion" in complemento:
            resultado = {}
            resultado.update(complemento)
            resultado.update({"@Tipo": "Percepcion"})
            resultado.update(
                {
                    "@Importe": "{:.2f}".format(
                        round(
                            float(complemento["@ImporteGravado"])
                            + float(complemento["@ImporteExento"]),
                            2,
                        )
                    )
                }
            )
            resultado["@TipoComplemento"] = resultado.pop("@TipoPercepcion")
            resultado["@Clave"] = resultado.pop("@Clave")
            ordered_percepcion = {key: resultado[key] for key in n_fields if key in resultado}
            ordered_percepcion = is_str_convert_to_float(
                old_dict=ordered_percepcion, keys_to_convert=keys_to_convert_to_float
            )
            final_data.update(ordered_percepcion)
            return tuple(final_data.values())
        elif "@TipoDeduccion" in complemento:
            resultado = {}
            resultado.update(commons_keys_dict)
            resultado.update(complemento)
            resultado.update({"@Tipo": "Deduccion"})
            resultado["@TipoComplemento"] = resultado.pop("@TipoDeduccion")
            resultado["@Clave"] = resultado.pop("@Clave")
            ordered_deduccion = {key: resultado[key] for key in n_fields if key in resultado}
            ordered_deduccion = is_str_convert_to_float(
                old_dict=ordered_deduccion, keys_to_convert=keys_to_convert_to_float
            )
            final_data.update(ordered_deduccion)
            return tuple(final_data.values())
        elif "@TipoOtroPago" in complemento:
            resultado = {}
            resultado.update(commons_keys_dict)

            combined_dict = {**complemento}
            combined_dict.pop("SubsidioAlEmpleo", None)
            resultado.update(combined_dict)
            resultado.update({"@Tipo": "Otro Pago"})
            resultado["@TipoComplemento"] = resultado.pop("@TipoOtroPago")
            resultado["@Clave"] = resultado.pop("@Clave")
            ordered_otro_pago_res = {key: resultado[key] for key in n_fields if key in resultado}
            ordered_otro_pago_res = is_str_convert_to_float(
                old_dict=ordered_otro_pago_res, keys_to_convert=keys_to_convert_to_float
            )
            final_data.update(ordered_otro_pago_res)
            return tuple(final_data.values())

    def _adjust_col_size(self, worksheet: Worksheet) -> None:
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = (
                length * self.COL_RESIZE_MAGIC_NUMBER
            )

    def _resume_to_worksheet(self, worksheet, resume, resume_type):
        fields_to_use = headers_by_resume_type[resume_type]

        headers = (self.resume_header_type,) + tuple(fields_to_use.values())
        worksheet.append(headers)
        keys = fields_to_use.keys()
        filtered = ("Periodo",) + tuple(resume[FILTERED][k] for k in keys)
        worksheet.append(filtered)
        exercise = ("Acumulado",) + tuple(resume[EXERCISE][k] for k in keys)
        worksheet.append(exercise)

    def _workbook_to_bytes(self, workbook: Workbook) -> bytes:
        xlsx_bytes = BytesIO()
        workbook.save(xlsx_bytes)
        xlsx_bytes.seek(0)
        return xlsx_bytes.read()

    def export(
        self,
        export_data: dict,
        cfdis: Iterable[CFDIORM],
        fields: Iterable[str],
        resume: CFDIResume,
        resume_type,
    ) -> bytes:
        work_book = Workbook()
        main_worksheet = work_book.active
        self._data_to_worksheet(main_worksheet, cfdis, fields, export_data=export_data)
        ws_totales = work_book.create_sheet("Totales")
        self._resume_to_worksheet(ws_totales, resume, resume_type)

        for ws in work_book.worksheets:
            self._adjust_col_size(ws)

        return self._workbook_to_bytes(work_book)

    def _add_additional_page(self, worksheet: Worksheet, data: list[list[str]]) -> None:
        for row in data:
            worksheet.append(row)

    def new_export(
        self,
        export_data: dict,
        cfdis: Iterable[CFDIORM],
        fields,
        additional_pages: dict[str, list[list[str]]] = None,
        export_type=None,
    ) -> bytes:
        work_book = Workbook()
        main_worksheet = work_book.active
        if export_data.get("iva") == "OpeConTer":
            self._data_to_worksheet(main_worksheet, cfdis, fields, export_data, title="Ingresos")
        else:
            self._data_to_worksheet(main_worksheet, cfdis, fields, export_data)

        additional_pages = additional_pages or {}
        for page_name, page_data in additional_pages.items():
            additional_page = work_book.create_sheet(page_name)
            self._add_additional_page(additional_page, page_data)

        for ws in work_book.worksheets:
            self._adjust_col_size(ws)

        return self._workbook_to_bytes(work_book)


def process_iterable(it: tuple[Any, ...]) -> str:
    it = tuple(it)
    if not it:
        return ""
    res_values = []
    for record in it:
        if isinstance(record, dict):
            values = tuple(record.values())
            if len(values) == 1 and is_primitive(values[0]):
                res_values.append(values[0])
    if res_values:
        return ", ".join(str(v) for v in res_values)
    return ", ".join(str(it))


def is_primitive(value: Any) -> bool:
    return isinstance(value, str | int | float | bool | date | datetime)


def is_str_convert_to_float(old_dict, keys_to_convert=None):
    if keys_to_convert is None:
        keys_to_convert = old_dict.keys()

    for key in keys_to_convert:
        if key in old_dict and old_dict[key] is not None:
            with contextlib.suppress(ValueError):
                old_dict[key] = float(old_dict[key])
    return old_dict

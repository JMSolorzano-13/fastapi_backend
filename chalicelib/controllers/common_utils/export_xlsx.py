import csv
from collections.abc import Callable
from datetime import UTC, date, datetime, timedelta
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Alignment
from sqlalchemy import BigInteger, Date, DateTime, Float, Integer, Numeric, SmallInteger
from sqlalchemy.orm import Session
from sqlalchemy.orm.query import Query

from chalicelib.boto3_clients import s3_client
from chalicelib.controllers.common_utils.s3.upload_public import upload_public
from chalicelib.new.cfdi_processor.domain.enums.cfdi_export_state import CfdiExportState
from chalicelib.new.config.infra import envars
from chalicelib.new.query.infra.copy_query import copy_query
from chalicelib.schema.models.tenant.cfdi_export import CfdiExport

INT_TYPES = (Integer, SmallInteger, BigInteger)
DECIMAL_TYPES = (Float, Numeric)
DATE_TYPES = Date
DATETIME_TYPES = DateTime

FORMATS: dict[tuple[type, ...] | type, tuple[str, Callable]] = {
    INT_TYPES: ("0", int),
    DECIMAL_TYPES: ("0.00", float),
    DATE_TYPES: ("YYYY-MM-DD", date.fromisoformat),
    DATETIME_TYPES: ("YYYY-MM-DD HH:MM:SS", datetime.fromisoformat),
}


def query_to_csv(query: Query):
    assert query.session
    session: Session = query.session
    string_io = StringIO()
    copy_query(session, query, string_io)
    string_io.seek(0)
    return csv.reader(string_io)


def get_column_letter(col_idx: int) -> str:
    """Convert a column index (1-based) to an Excel column letter."""
    letter = ""
    while col_idx > 0:
        col_idx -= 1
        letter = chr(col_idx % 26 + 65) + letter
        col_idx //= 26
    return letter


def query_to_xlsx(query: Query) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet

    def load_data():
        csv_data = query_to_csv(query)
        for row in csv_data:
            worksheet.append(row)

    def get_column_formats():
        custom_styles: dict[str, tuple[str, Callable]] = {}  # 'A', 'B', 'C', 'D'
        for i, col in enumerate(query.column_descriptions):
            for sql_type, (fmt, py_type) in FORMATS.items():
                if isinstance(col["type"], sql_type):
                    custom_styles[get_column_letter(i + 1)] = (fmt, py_type)
                    break
        return custom_styles

    def set_column_formats():
        custom_styles = get_column_formats()
        for col, (fmt, py_type) in custom_styles.items():
            for cell in worksheet[col]:
                if cell.row == 1:  # Saltar cabecera
                    continue
                cell.number_format = fmt
                try:
                    cell.value = py_type(cell.value)
                except ValueError:
                    print(
                        f"Error converting cell {cell.coordinate} to {py_type.__name__}, value: {cell.value}"  # noqa: E501
                    )
                    pass
                if fmt == "0.00":
                    cell.alignment = Alignment(horizontal="right")

    def adjust_col_size():
        first_row_to_analyze = 1
        max_row_to_analyze = 10
        max_width = 100
        COL_RESIZE_MAGIC_NUMBER = 1.1
        for column_cells in worksheet.iter_cols(
            min_row=first_row_to_analyze, max_row=max_row_to_analyze
        ):
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                length * COL_RESIZE_MAGIC_NUMBER, max_width
            )

    load_data()
    set_column_formats()
    adjust_col_size()

    return workbook


def _workbook_to_bytes(workbook: Workbook) -> bytes:
    xlsx_bytes = BytesIO()
    workbook.save(xlsx_bytes)
    xlsx_bytes.seek(0)
    return xlsx_bytes.read()


def generic_xlsx_export(
    file_name: str,
    query: Query,
) -> CfdiExport:
    xlsx = query_to_xlsx(query)

    query_str = str(query.statement.compile(compile_kwargs={"literal_binds": True}))

    xlsx_bytes = _workbook_to_bytes(xlsx)
    expiration_delta = timedelta(days=1)
    expiration_date = datetime.now(tz=UTC) + expiration_delta
    url = upload_public(
        s3_client=s3_client(),
        data=xlsx_bytes,
        bucket=envars.S3_EXPORT,
        key=file_name,
        expiration_delta=expiration_delta,
    )
    export_metadata = CfdiExport(
        url=url,
        state=CfdiExportState.TO_DOWNLOAD,
        expiration_date=expiration_date,
        start="",
        end="",
        cfdi_type="",
        download_type="",
        format="XLSX",
        displayed_name="",
        file_name=file_name,
        domain=query_str,
    )
    # export_repo_s3 = ExportRepositoryS3()
    with open("/tmp/export.xlsx", "wb") as f:
        xlsx.save(f)
    # export_repo_s3.save(xlsx_bytes, export_metadata, {"file_name": file_name})
    return export_metadata

import json
from io import BytesIO
from uuid import uuid4

from chalice.test import Client
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from chalicelib.controllers.cfdi import CFDIController
from chalicelib.controllers.enums import ResumeType
from chalicelib.new.cfdi_processor.domain.xlsx_exporter import XLSXExporter
from chalicelib.schema.models import Company
from chalicelib.schema.models.tenant.cfdi import CFDI
from chalicelib.schema.models.tenant.nomina import Nomina
from chalicelib.schema.models.tenant.payment import Payment


def test_selected_xlsx_export_payments(
    client_authenticated: Client, company: Company, company_session: Session
):
    ingreso = CFDI.demo(
        company_identifier=company.identifier,
        is_issued=True,
        UUID=str(uuid4()),
        TipoDeComprobante="P",
        PaymentDate="2020-12-15",
    )
    payment = Payment(
        company_identifier=company.identifier,
        is_issued=True,
        uuid_origin=ingreso.UUID,
        index=1,
        FechaPago="2020-12-15",
        FormaDePagoP="01",
        MonedaP="MXN",
        Monto=1000,
        TipoCambioP=1,
        NumOperacion="1234",
        RfcEmisorCtaOrd="AAAAA010101AAA",
        NomBancoOrdExt="Banco de Mexico",
        CtaOrdenante="1234",
        RfcEmisorCtaBen="AAAAA010101AAA",
    )

    company_session.add(ingreso)
    company_session.add(payment)

    company_session.commit()

    response = client_authenticated.http.post(
        "/CFDI/export",
        body=json.dumps(
            {
                "domain": [
                    ["UUID", "in", [ingreso.UUID]],
                    ["company_identifier", "=", company.identifier],
                ],
                "export_data": {"file_name": "Pagos", "type": ""},
                "fields": [
                    "UUID",
                    "payments.Monto",
                ],
                "format": "XLSX",
            }
        ),
        headers={
            "Content-Type": "application/json",
        },
    )
    assert response.status_code == 200


def test_selected_xlsx_export_by_function_payments(
    company_session: Session,
    company: Company,
):
    ingreso = CFDI.demo(
        company_identifier=company.identifier,
        is_issued=True,
        UUID=str(uuid4()),
        TipoDeComprobante="P",
        PaymentDate="2020-12-15",
    )
    payment = Payment(
        company_identifier=company.identifier,
        is_issued=True,
        uuid_origin=ingreso.UUID,
        index=1,
        FechaPago="2020-12-15",
        FormaDePagoP="01",
        MonedaP="MXN",
        Monto=1000,
        TipoCambioP=1,
        NumOperacion="1234",
        RfcEmisorCtaOrd="AAAAA010101AAA",
        NomBancoOrdExt="Banco de Mexico",
        CtaOrdenante="1234",
        RfcEmisorCtaBen="AAAAA010101AAA",
    )

    company_session.add(ingreso)
    company_session.add(payment)

    company_session.commit()

    search_attrs = {
        "fuzzy_search": None,
        "fields": [
            "payments.NomBancoOrdExt",
            "payments.c_forma_pago.name",
        ],
        "domain": [
            ["UUID", "in", [ingreso.UUID]],
            ["company_identifier", "=", company.identifier],
        ],
        "order_by": None,
        "limit": None,
        "offset": None,
        "active": True,
    }

    controller = CFDIController()

    query = controller._search(
        **search_attrs,
        lazzy=True,
        session=company_session,
    )
    resume_type = ResumeType["P"]

    resume_export = controller.resume(
        domain=search_attrs["domain"],
        session=company_session,
        context={},
        resume_type=resume_type,
        fuzzy_search=None,
        fields=search_attrs["fields"],
    )

    export_data = {"file_name": "CFO1010219Z8_Recibidos_pagos_Dic2025", "type": ""}

    xlsx_exporter = XLSXExporter()
    data_bytes = xlsx_exporter.export(
        export_data,
        query,
        fields=search_attrs["fields"],
        resume=resume_export,
        resume_type=resume_type,
    )

    wb = load_workbook(filename=BytesIO(data_bytes))

    sheet = wb.active

    assert sheet.cell(row=1, column=1).value == "Banco ordenante"
    assert sheet.cell(row=2, column=1).value == "Banco de Mexico"


def test_selected_xlsx_export_by_function_payroll(
    company_session: Session,
    company: Company,
):
    payroll = CFDI.demo(
        company_identifier=company.identifier,
        is_issued=True,
        UUID=str(uuid4()),
        TipoDeComprobante="N",
        PaymentDate="2020-12-15",
    )

    related_payroll = Nomina(
        company_identifier=company.identifier,
        cfdi_uuid=payroll.UUID,
        Version="1.2",
        TipoNomina="O",
        FechaPago="2020-12-15",
        FechaInicialPago="2020-12-01",
        FechaFinalPago="2020-12-15",
        NumDiasPagados=15,
        TotalPercepciones=1000,
        TotalDeducciones=100,
        TotalOtrosPagos=0,
        ReceptorCurp="AAAA010101HDFAAA01",
        ReceptorNumSeguridadSocial="12345678901",
        ReceptorFechaInicioRelLaboral="2020-12-01",
        ReceptorAntigüedad="P1Y",
        ReceptorTipoContrato="01",
        ReceptorTipoRegimen="02",
        ReceptorNumEmpleado="1234",
        ReceptorDepartamento="Ventas",
        ReceptorPuesto="Vendedor",
        ReceptorPeriodicidadPago="02",
        ReceptorClaveEntFed="CMX",
    )
    company_session.add(payroll)
    company_session.add(related_payroll)
    company_session.commit()

    search_attrs = {
        "fuzzy_search": None,
        "fields": [
            "nomina.ReceptorTipoRegimen",
            "nomina.FechaPago",
        ],
        "domain": [
            ["UUID", "in", [payroll.UUID]],
            ["company_identifier", "=", company.identifier],
        ],
        "order_by": None,
        "limit": None,
        "offset": None,
        "active": True,
    }

    controller = CFDIController()

    query = controller._search(  # pylint: disable=protected-access
        **search_attrs,
        lazzy=True,
        session=company_session,
    )
    resume_type = ResumeType["N"]

    resume_export = controller.resume(
        domain=search_attrs["domain"],
        session=company_session,
        context={},
        resume_type=resume_type,
        fuzzy_search=None,
        fields=search_attrs["fields"],
    )

    export_data = {"file_name": "Nominas", "type": ""}

    xlsx_exporter = XLSXExporter()
    data_bytes = xlsx_exporter.export(
        export_data,
        query,
        fields=search_attrs["fields"],
        resume=resume_export,
        resume_type=resume_type,
    )

    wb = load_workbook(filename=BytesIO(data_bytes))

    sheet = wb.active

    assert sheet.cell(row=1, column=1).value == "Tipo régimen"
    assert sheet.cell(row=2, column=1).value == "02"

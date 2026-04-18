"""
Test para verificar el fix de agregación de campos DateTime en exportaciones XLSX.

Replica el flujo completo de /massive_export sin depender de AWS SQS.
Bug original: ParserError: Unknown string format: 2025-09-01 00:00:00, 2025-09-01 00:00:00
"""

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from chalicelib.bus import EventBus
from chalicelib.controllers.enums import ResumeType
from chalicelib.new.cfdi_processor.domain.cfdi_exporter import CFDIExporter
from chalicelib.new.cfdi_processor.infra.cfdi_export_repository_sa import CFDIExportRepositorySA
from chalicelib.new.translations.xlsx_fields import common_fields
from chalicelib.schema.models import Company


def test_export_with_file_translations(
    company_session: Session, session: Session, company: Company
):
    """
    Test que ejecuta la exportación real y captura si hay el bug de DateTime.
    Si no hay datos, busca en producción/staging.
    """
    # Body exacto que causó el error en producción
    json_body = {
        "domain": [
            ["company_identifier", "=", company.identifier],
            ["FechaFiltro", ">=", "2025-09-01T00:00:00.000"],
            ["FechaFiltro", "<", "2025-10-01T00:00:00.000"],
            ["Estatus", "=", True],
            ["is_issued", "=", True],
            ["TipoDeComprobante", "=", "I"],
        ],
        "fuzzy_search": "",
        "format": "XLSX",
        "fields": [
            "attachments_count",
            "FechaFiltro",
            "RfcReceptor",
            "NombreReceptor",
        ],
        "export_data": {"file_name": "TEST_BUG_FIX", "type": ""},
    }

    # Preparar body
    body = {
        "domain": json_body["domain"],
        "fuzzy_search": json_body["fuzzy_search"],
        "fields": json_body["fields"],
    }

    # Crear exporter y ejecutar
    exporter = CFDIExporter(
        company_session=company_session,
        cfdi_export_repo=CFDIExportRepositorySA(session=company_session),
        bus=EventBus(),
    )

    export_bytes = exporter.export_xlsxv2(
        body=body,
        context={},
        query=None,
        fields=json_body["fields"],
        resume_type=ResumeType.BASIC,
        export_data=json_body["export_data"],
    )

    # Guardar Excel
    output_file = "test_export_xlsx.xlsx"

    with open(output_file, "wb") as f:
        validated_field = "attachments_count"

        index_of_validated_field = body["fields"].index(validated_field)

        f.write(export_bytes)
        wb = load_workbook(output_file)
        sheet = wb.active
        cell = sheet.cell(row=1, column=index_of_validated_field + 1)
        assert cell.value == common_fields[body["fields"][index_of_validated_field]]

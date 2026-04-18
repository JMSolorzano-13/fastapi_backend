from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from chalicelib.controllers.cfdi import CFDIController
from chalicelib.controllers.docto_relacionado import DoctoRelacionadoController
from chalicelib.new.isr_deducciones import (
    build_total_deducciones_cfdi_query,
    build_total_deducciones_pagos_query,
    calcular_totales_nomina_data,
    get_isr_percentage,
)
from chalicelib.new.isr_exportaciones import (
    ISR_CFDI,
    ISR_DOCTORELACIONADO,
    _export_isr_generic,
    create_export_record,
    export_total_isr_page,
)
from chalicelib.schema.models.company import Company
from chalicelib.schema.models.tenant import CFDI


def setup_test_data(company_session, count=5, **kwargs):
    """Setup test CFDI data"""
    defaults = {"Moneda": "MXN", "Estatus": True, "is_issued": True}
    defaults.update(kwargs)

    company_session.add_all(CFDI.demo(**defaults) for _ in range(count))
    company_session.flush()


def test_isr_export_cfdi_endpoint(
    company_session: Session,
    session: Session,
    company: Company,
):
    setup_test_data(company_session, count=5, Moneda="MXN", TipoDeComprobante="I")
    company_session.commit()

    export_payload = {
        "period": "2024-01-01",
        "yearly": False,
        "issued": True,
        "company_identifier": company.identifier,
        "export_data": {"file_name": "test_isr_endpoint_export.xlsx"},
        "displayed_name": "Test ISR Endpoint Export",
        "domain": [
            ["company_identifier", "=", company.identifier],
            ["Estatus", "=", True],
            ["is_issued", "=", True],
            ["TipoDeComprobante", "=", "I"],
            ["is_issued", "=", True],
        ],
        "fields": {
            "UUID": "UUID",
            "TipoDeComprobante": "Tipo de Comprobante",
            "Total": "Total",
            "Fecha": "Fecha",
        },
        "total_cfdi": {
            "domain_totales": [
                ["company_identifier", "=", company.identifier],
                ["TipoDeComprobante", "=", "I"],
                ["Estatus", "=", True],
            ],
            "fields_totales": ["SubTotal", "DescuentoMXN", "NetoMXN", "RetencionesISRMXN"],
        },
    }

    # Generar XLSX con hojas CFDI y Totales
    workbook_bytes = _export_isr_generic(
        company_session=company_session,
        export_payload=export_payload,
        controller_class=CFDIController,
        column_mapping=ISR_CFDI,
        total_key="total_cfdi",
        build_total_query_func=build_total_deducciones_cfdi_query,
    )

    # ASSERTS
    assert workbook_bytes is not None
    assert isinstance(workbook_bytes, bytes)

    # Load workbook from bytes to verify structure
    workbook = load_workbook(BytesIO(workbook_bytes))
    assert len(workbook.sheetnames) == 2
    assert workbook["CFDI"].max_row >= 1
    assert workbook["Totales"].max_row == 2

    # CREATE EXPORT REQUEST
    export_request = create_export_record(company_session, export_payload)
    assert export_request.identifier is not None


def test_isr_export_totales_endpoint(
    company_session: Session,
    session: Session,
    company: Company,
):
    setup_test_data(company_session, count=5, Moneda="MXN", TipoDeComprobante="N")
    company_session.commit()

    company = session.query(Company).filter_by(identifier=company.identifier).first()

    export_payload = {
        "period": "2024-01-01",
        "yearly": False,
        "issued": True,
        "company_identifier": company.identifier,
        "export_data": {"file_name": "test_isr_totales_export.xlsx"},
        "displayed_name": "Test ISR Totales Export",
        "total_nomina": {"domain_totales": [["company_identifier", "=", company.identifier]]},
    }

    # ENDPOINT
    period_date = datetime.fromisoformat(export_payload.get("period")).date()

    # DATA
    isr_data = calcular_totales_nomina_data(company_session, session, company, period_date)

    # XLSX
    workbook_bytes = export_total_isr_page(isr_data)

    # ASSERTS
    assert workbook_bytes is not None
    assert isinstance(workbook_bytes, bytes)

    # Load workbook from bytes to verify structure
    workbook = load_workbook(BytesIO(workbook_bytes))
    assert len(workbook.sheetnames) == 1
    assert "Totales" in workbook.sheetnames

    # CREATE EXPORT REQUEST
    export_request = create_export_record(company_session, export_payload)
    assert export_request.identifier is not None


def test_isr_export_totales_percentage(session: Session, company: Company):
    """
    Verifica que get_isr_percentage() convierta correctamente floats a Decimal.

    Valida que valores como 0.47 y 0.53 se conviertan exactamente a Decimal("0.47")
    y Decimal("0.53"), sin pérdida de precisión, usando Decimal(str(...)).
    """
    company_obj = session.query(Company).filter_by(identifier=company.identifier).first()

    # Caso 1: 0.47 debe ser exactamente Decimal("0.47")
    company_obj.data = {"isr_percentage": 0.47}
    session.flush()

    isr_pct_47 = get_isr_percentage(company_obj)
    assert isinstance(isr_pct_47, Decimal), "Debe retornar un Decimal"
    assert isr_pct_47 == Decimal("0.47"), "Debe ser exactamente 0.47"
    assert int(isr_pct_47 * 100) == 47, "47% debe mostrarse como 47, no 46"

    # Caso 2: 0.53 debe ser exactamente Decimal("0.53")
    company_obj.data = {"isr_percentage": 0.53}
    session.flush()

    isr_pct_53 = get_isr_percentage(company_obj)
    assert isinstance(isr_pct_53, Decimal), "Debe retornar un Decimal"
    assert isr_pct_53 == Decimal("0.53"), "Debe ser exactamente 0.53"
    assert int(isr_pct_53 * 100) == 53, "53% debe mostrarse como 53"


def test_isr_export_pagos_endpoint(
    company_session: Session,
    session: Session,
    company: Company,
):
    setup_test_data(company_session, count=3, Moneda="MXN", TipoDeComprobante="I")
    company_session.commit()

    export_payload = {
        "period": "2024-01-01",
        "yearly": False,
        "issued": True,
        "company_identifier": company.identifier,
        "export_data": {"file_name": "test_isr_pagos_export.xlsx"},
        "displayed_name": "Test ISR Pagos Export",
        "domain": [
            ["company_identifier", "=", company.identifier],
            ["is_issued", "=", True],
        ],
        "fields": {
            "UUID": "UUID",
            "UUID_related": "UUID Relacionado",
        },
        "total_pagos": {
            "domain_totales": [
                ["company_identifier", "=", company.identifier],
                ["is_issued", "=", True],
            ],
            "fields_totales": [
                "BaseIVA16",
                "BaseIVA8",
                "BaseIVA0",
                "BaseIVAExento",
                "Neto",
                "RetencionesISR",
            ],
        },
    }

    # Generar XLSX con hojas CFDI y Totales
    workbook_bytes = _export_isr_generic(
        company_session=company_session,
        export_payload=export_payload,
        controller_class=DoctoRelacionadoController,
        column_mapping=ISR_DOCTORELACIONADO,
        total_key="total_pagos",
        build_total_query_func=build_total_deducciones_pagos_query,
    )

    # ASSERTS
    assert workbook_bytes is not None
    assert isinstance(workbook_bytes, bytes)

    # Load workbook from bytes to verify structure
    workbook = load_workbook(BytesIO(workbook_bytes))
    assert len(workbook.sheetnames) == 2
    assert workbook["Totales"].max_row == 2

    # CREATE EXPORT REQUEST
    export_request = create_export_record(company_session, export_payload)
    assert export_request.identifier is not None

"""
Prueba de regresión: GroupingError al exportar documentos relacionados de Pagos.

El bug: is_aggregated=True para el caso 'doctos' causaba que func.max() se
aplicara sobre la query de pago_docs_relacionados (sin GROUP BY).
"""

import io
from datetime import datetime
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from chalicelib.controllers.cfdi import CFDIController
from chalicelib.controllers.enums import ResumeType
from chalicelib.new.cfdi_processor.domain.cfdi_exporter import CFDIExporter
from chalicelib.new.cfdi_processor.infra.cfdi_export_repository_sa import CFDIExportRepositorySA
from chalicelib.schema.models import Company
from chalicelib.schema.models.tenant import DoctoRelacionado
from chalicelib.schema.models.tenant.cfdi import CFDI

FECHA_PAGO = datetime(2025, 1, 21, 12, 48, 41)


def _insert_pagos_con_demo(company_session: Session, company: Company) -> None:
    """
    Crea 3 CFDIs de pago (CFDI.demo tipo P) + 3 DoctoRelacionado.demo
    apuntando a cada pago.
    """
    for i in range(3):
        cfdi_pago = CFDI.demo(
            company_identifier=company.identifier,
            Fecha=FECHA_PAGO,
            FechaFiltro=FECHA_PAGO,
            PaymentDate=FECHA_PAGO,
            TipoDeComprobante="P",
            is_issued=False,
            Estatus=True,
            Moneda="XXX",
            Total=Decimal("0.00"),
            SubTotal=Decimal("0.00"),
            Serie="PAG",
            Folio=str(5822 + i),
        )
        company_session.add(cfdi_pago)
        company_session.flush()

        company_session.add(
            DoctoRelacionado.demo(
                company_identifier=company.identifier,
                is_issued=False,
                UUID=cfdi_pago.UUID,  # → apunta al CFDI de pago
                FechaPago=FECHA_PAGO,
                NumParcialidad=i + 1,
                Estatus=True,
                active=True,
            )
        )

    company_session.commit()


def test_export_pago_docs_relacionados_export(company_session: Session, company: Company):
    """
    Regresión: export_xlsxv2 con export_data["type"]="doctos" no debe lanzar
    GroupingError de PostgreSQL. Usa CFDI.demo + DoctoRelacionado.demo.
    """
    _insert_pagos_con_demo(company_session, company)

    body = {
        "domain": [
            ["company_identifier", "=", company.identifier],
            ["FechaFiltro", ">=", "2025-01-01T00:00:00.000"],
            ["FechaFiltro", "<", "2026-01-21T00:00:00.000"],
            ["Estatus", "=", True],
            ["is_issued", "=", False],
            ["TipoDeComprobante", "=", "P"],
        ],
        "fuzzy_search": "",
        "format": "XLSX",
        "limit": None,
        "offset": None,
        "order_by": None,
    }

    fields = [
        "Fecha",
        "UUID",
        "Serie",
        "Folio",
        "RfcReceptor",
        "NombreReceptor",
        "pays.UUID_related",
        "UsoCFDIReceptor",
        "pays.MonedaDR",
        "pays.NumParcialidad",
        "pays.ImpPagado",
        "pays.cfdi_origin.Fecha",
    ]

    export_data = {
        "file_name": "CFO1010219Z8_pagosDocsRelacionados_emitidos_2025",
        "type": "doctos",
    }

    search_body = {k: body[k] for k in ("domain", "fuzzy_search", "limit", "offset", "order_by")}
    query = CFDIController._search(
        **search_body, fields=fields, session=company_session, lazzy=True
    )

    exporter = CFDIExporter(
        company_session,
        cfdi_export_repo=CFDIExportRepositorySA(session=company_session),
    )

    # Antes del fix lanzaba GroupingError de PostgreSQL
    file = exporter.export_xlsxv2(
        body=body,
        query=query,
        fields=fields,
        resume_type=ResumeType.P,
        export_data=export_data,
        context=None,
    )

    filename = "test_pago_docs_relacionados.xlsx"
    with open(filename, "wb") as f:
        f.write(file)

    assert file is not None and len(file) > 0

    wb = load_workbook(io.BytesIO(file), data_only=True)
    ws = wb["CFDI"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data_rows = [
        r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)
    ]

    assert len(data_rows) == 3, f"Se esperaban 3 filas, se obtuvieron {len(data_rows)}"
    assert "DR - Numero de parcialidad" in headers

    print("✓ Test Pagos docs relacionados pasó correctamente")
    print(f"  📄 Excel generado: {filename} ({len(file)} bytes)")
    print(f"  📊 Filas de datos: {len(data_rows)}")

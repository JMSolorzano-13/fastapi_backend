import io

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from chalicelib.controllers.cfdi import CFDIController
from chalicelib.controllers.enums import ResumeType
from chalicelib.new.cfdi_processor.domain.query_to_export.pago_docs_relacionados import (
    pagos_column_types,
    query_pago_docs_relacionados,
)
from chalicelib.new.cfdi_processor.domain.xlsx_v2 import ExportV2
from chalicelib.new.query.domain.xml_processor import XMLProcessor
from chalicelib.new.query.infra.cfdi_repository_sa import CFDIRepositorySA
from chalicelib.schema.models.company import Company
from tests.load_data.test_company_load import read_files_from_directory

EXPECTED_HEADERS = [
    "Fecha de pago",
    "Fecha de emisión",
    "UUID",
    "Serie",
    "Folio",
    # IS_ISSUED TRUE/FALSE
    # "RFC receptor",
    # "Receptor",
    "RFC emisor",
    "Emisor",
    "Forma de pago",
    "Moneda de pago",
    "DR - Fecha de emisión",
    "DR - Serie",
    "DR - Folio",
    "DR - UUID",
    "DR - Uso de CFDI",
    "DR - Objeto de impuesto",
    "DR - Moneda",
    "DR - Equivalencia",
    "DR - Numero de parcialidad",
    "DR - Importe pagado",
    "DR - Importe pagado MXN",
    "DR - Base IVA 16 %",
    "DR - Base IVA 8 %",
    "DR - Base IVA 0 %",
    "DR - Base IVA Exento",
    "DR - IVA 16 %",
    "DR - IVA 8 %",
    "DR - IVA total",
    "DR - Base IEPS",
    "DR - Factor IEPS",
    "DR - Tasa o cuota IEPS",
    "DR - IEPS",
    "DR - Retenciónes ISR",
    "DR - Retenciónes IVA",
]

XMLS_PATH = "tests/chalicelib/new/cfdi_processor/xml_docs_related"


def insert_specific_cfdi(c_session: Session, company: Company):
    cfdi_repo = CFDIRepositorySA(session=c_session)
    xml_content = read_files_from_directory(XMLS_PATH)

    xmlProcessor = XMLProcessor(cfdi_repo=cfdi_repo, xml_repo=None, company_session=c_session)

    xmlProcessor.process_xml_files(
        company_identifier=company.identifier,
        xmls_contents=xml_content,
        rfc=company.rfc,
    )


def test_export_with_docs_related(company_session: Session, company: Company):
    insert_specific_cfdi(company_session, company)
    body = {
        "domain": [
            ["FechaFiltro", ">=", "2025-08-01T00:00:00.000"],
            ["FechaFiltro", "<", "2025-09-01T00:00:00.000"],
            ["Estatus", "=", True],
            ["is_issued", "=", False],
            ["TipoDeComprobante", "=", "P"],
        ],
        "fuzzy_search": "",
        "format": "XLSX",
        "fields": [
            "Fecha",
            "UUID",
            "Serie",
            "Folio",
            "RfcReceptor",
            "NombreReceptor",
            "pays.UUID_related",
            "UsoCFDIReceptor",
            "pays.MonedaDR",
            "pays.NumParcialidad",
            "pays.ImpPagado",
            "pays.cfdi_related.Fecha",
        ],
        "TipoDeComprobante": "P",
        "export_type": [
            "Fecha",
            "UUID",
            "Serie",
            "Folio",
            "RfcReceptor",
            "NombreReceptor",
            "pays.UUID_related",
            "UsoCFDIReceptor",
            "pays.MonedaDR",
            "pays.NumParcialidad",
            "pays.ImpPagado",
            "pays.cfdi_related.Fecha",
        ],
        "export_data": {
            "file_name": "PGD1009214W0_pagosDocsRelacionados_emitidos_Desde1Ene2024_Hasta31Ene2024",
            "type": "doctos",
        },
    }
    query = query_pago_docs_relacionados(company_session, body["domain"])
    column_types_override = pagos_column_types
    query = CFDIController.apply_domain(
        query=query,
        domain=body["domain"],
        fuzzy_search=body["fuzzy_search"],
        session=company_session,
    )

    export_v2 = ExportV2(company_session)

    xlsx_bytes = export_v2.export(query, body, [], ResumeType.P, column_types_override)
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 100
    assert xlsx_bytes.startswith(b"PK\x03\x04")

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    assert wb.sheetnames[0] == "CFDI"
    ws = wb["CFDI"]
    headers_in_sheet = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert set(EXPECTED_HEADERS).issubset(headers_in_sheet)

    col_idx_importe = headers_in_sheet.index("DR - Importe pagado")
    col_idx_importe_mxn = headers_in_sheet.index("DR - Importe pagado MXN")

    # leer primera fila de datos (fila 2 en Excel)
    first_data_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    importe_dr = first_data_row[col_idx_importe]
    importe_mxn_dr = first_data_row[col_idx_importe_mxn]

    assert importe_dr == 250.0
    assert importe_mxn_dr == 4550.0

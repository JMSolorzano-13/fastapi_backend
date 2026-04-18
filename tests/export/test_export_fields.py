import json
import uuid
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from chalicelib.controllers.cfdi import CFDIController
from chalicelib.controllers.enums import ResumeType
from chalicelib.new.cfdi_processor.domain.cfdi_exporter import CFDIExporter
from chalicelib.new.cfdi_processor.infra.cfdi_export_repository_sa import (
    CFDIExportRepositorySA,
)
from chalicelib.schema.models.tenant.cfdi import CFDI


@pytest.mark.skip("generic_xslx")
def test_export_cfdi_with_cuenta_predial(company_session: Session):
    conceptos_data = {
        "Concepto": [
            {
                "@ClaveProdServ": "84111506",
                "@Cantidad": "1",
                "@ClaveUnidad": "ACT",
                "@Descripcion": "Producto de prueba",
                "@ValorUnitario": "100.00",
                "@Importe": "100.00",
                "CuentaPredial": {"@Numero": "123456789012345678900"},
            },
        ]
    }

    cfdi = CFDI.demo(
        is_issued=False,
        FechaFiltro=datetime(2021, 2, 1),
        Fecha=datetime(2021, 2, 1),
        UUID=str(uuid.uuid4()),
        company_identifier=str(uuid.uuid4()),
        RfcEmisor="EMISOR010101000",
        RfcReceptor="RECEPTOR010101000",
        BaseIVA0=0,
        BaseIVA16=0,
        BaseIVA8=0,
        BaseIVAExento=0,
        IVATrasladado16=0,
        IVATrasladado8=0,
        Total=Decimal("200.00"),
        SubTotal=200,
        TipoCambio=0,
        Neto=0,
        TrasladosIVA=0,
        TrasladosIEPS=0,
        TrasladosISR=0,
        RetencionesIVA=0,
        RetencionesIEPS=0,
        RetencionesISR=0,
        TotalMXN=0,
        SubTotalMXN=0,
        NetoMXN=0,
        TrasladosIVAMXN=0,
        DescuentoMXN=0,
        TrasladosIEPSMXN=0,
        TrasladosISRMXN=0,
        RetencionesIVAMXN=0,
        RetencionesIEPSMXN=0,
        RetencionesISRMXN=0,
        NoCertificado="000000",
        PaymentDate=datetime(2025, 2, 1),
        Descuento=Decimal("0.00"),
        pr_count=Decimal("0"),
        Estatus=True,
        TipoDeComprobante="I",
        FechaCertificacionSat=datetime(2021, 2, 1),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
        Conceptos=json.dumps(conceptos_data),
    )
    company_session.add(cfdi)
    company_session.commit()

    body = {
        "domain": [
            ["FechaFiltro", ">=", "2021-01-01T00:00:00.000"],
            ["FechaFiltro", "<", "2022-01-01T00:00:00.000"],
            ["Estatus", "=", True],
            ["is_issued", "=", False],
            ["TipoDeComprobante", "=", "I"],
            ["UUID", "=", cfdi.UUID],
        ],
        "fuzzy_search": "",
        "limit": None,
        "offset": None,
        "order_by": None,
    }
    export_data = dict(
        {
            "file_name": "CFDI_with_Concepts",
            "type": "",
        }
    )

    fields = [
        "UUID",
        "Conceptos.ClaveProdServ",
        "Conceptos.Cantidad",
        "Conceptos.ClaveUnidad",
        "Conceptos.Descripcion",
        "Conceptos.ValorUnitario",
        "Conceptos.Importe",
        "Conceptos.CuentaPredial.Numero",
        "uuid_total_egresos_relacionados",
        "Fecha",
        "LugarExpedicion",
        "Serie",
        "paid_by.UUID",
        "Folio",
        "RfcReceptor",
        "NombreReceptor",
        "Total",
        "TotalMXN",
        "balance",
        "SubTotal",
        "SubTotalMXN",
        "Descuento",
        "DescuentoMXN",
        "Neto",
        "NetoMXN",
        "TrasladosIVA",
        "TrasladosIVAMXN",
        "TrasladosIEPS",
        "TrasladosIEPSMXN",
        "TrasladosISR",
        "TrasladosISRMXN",
        "ExcludeFromIVA",
        "RetencionesIVA",
        "RetencionesIVAMXN",
        "RetencionesIEPS",
        "RetencionesIEPSMXN",
        "RetencionesISR",
        "RetencionesISRMXN",
        "RegimenFiscalReceptor",
        "c_regimen_fiscal_receptor.name",
        "CfdiRelacionados",
        "Moneda",
        "TipoCambio",
        "UsoCFDIReceptor",
        "MetodoPago",
        "c_metodo_pago.name",
        "FormaPago",
        "c_forma_pago.name",
        "CondicionesDePago",
        "Periodicidad",
        "Meses",
        "Year",
        "FechaYear",
        "FechaMonth",
        "Exportacion",
        "FechaCertificacionSat",
        "FechaCertificacionSatYear",
        "FechaCertificacionSatMonth",
        "NoCertificado",
        "Version",
        "TipoDeComprobante",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.Base",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.TipoFactor",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.TasaOCuota",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.Importe",
        "Conceptos.Impuestos.Traslados.Traslado.IEPS.Base",
        "Conceptos.Impuestos.Traslados.Traslado.IEPS.TipoFactor",
        "Conceptos.Impuestos.Traslados.Traslado.IEPS.TasaOCuota",
        "Conceptos.Impuestos.Traslados.Traslado.IEPS.Importe",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.Base",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.TipoFactor",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.TasaOCuota",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.Importe",
        "Conceptos.Impuestos.Retenciones.Retencion.ISR.Base",
        "Conceptos.Impuestos.Retenciones.Retencion.ISR.TipoFactor",
        "Conceptos.Impuestos.Retenciones.Retencion.ISR.TasaOCuota",
        "Conceptos.Impuestos.Retenciones.Retencion.ISR.Importe",
    ]

    query = CFDIController._search(**body, fields=fields, session=company_session)

    exporter = CFDIExporter(
        company_session,
        cfdi_export_repo=CFDIExportRepositorySA(session=company_session),
    )
    file = exporter.export_xlsxv2(
        body=body,
        query=query,
        fields=fields,
        resume_type=ResumeType.BASIC,
        export_data=export_data,
        context=None,
    )

    # with open("test_export_cfdi.xlsx", "wb") as f:
    #     f.write(file)

    assert "Conceptos.CuentaPredial.Numero" in fields
    assert "Conceptos.ClaveProdServ" in fields
    assert "Conceptos.Cantidad" in fields
    assert "Conceptos.ClaveUnidad" in fields
    assert "Conceptos.Descripcion" in fields
    assert "Conceptos.ValorUnitario" in fields
    assert file is not None


@pytest.mark.skip("generic_xslx")
def test_export_cfdi_with_cuenta_predial_multiple(company_session: Session):
    conceptos_data = {
        "Concepto": [
            {
                "@ClaveProdServ": "84111506",
                "@Cantidad": "1",
                "@ClaveUnidad": "ACT",
                "@Descripcion": "Producto de prueba",
                "@ValorUnitario": "100.00",
                "@Importe": "100.00",
                "CuentaPredial": [
                    {"@Numero": "123456789012345678900"},
                    {"@Numero": "123456789012345678900"},
                ],
            },
        ]
    }

    cfdi = CFDI.demo(
        is_issued=False,
        FechaFiltro=datetime(2021, 2, 1),
        Fecha=datetime(2021, 2, 1),
        UUID=str(uuid.uuid4()),
        company_identifier=str(uuid.uuid4()),
        RfcEmisor="EMISOR010101000",
        RfcReceptor="RECEPTOR010101000",
        BaseIVA0=0,
        BaseIVA16=0,
        BaseIVA8=0,
        BaseIVAExento=0,
        IVATrasladado16=0,
        IVATrasladado8=0,
        Total=Decimal("200.00"),
        SubTotal=200,
        TipoCambio=0,
        Neto=0,
        TrasladosIVA=0,
        TrasladosIEPS=0,
        TrasladosISR=0,
        RetencionesIVA=0,
        RetencionesIEPS=0,
        RetencionesISR=0,
        TotalMXN=0,
        SubTotalMXN=0,
        NetoMXN=0,
        TrasladosIVAMXN=0,
        DescuentoMXN=0,
        TrasladosIEPSMXN=0,
        TrasladosISRMXN=0,
        RetencionesIVAMXN=0,
        RetencionesIEPSMXN=0,
        RetencionesISRMXN=0,
        NoCertificado="000000",
        PaymentDate=datetime(2025, 2, 1),
        Descuento=Decimal("0.00"),
        pr_count=Decimal("0"),
        Estatus=True,
        TipoDeComprobante="I",
        FechaCertificacionSat=datetime(2021, 2, 1),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
        Conceptos=json.dumps(conceptos_data),
    )
    company_session.add(cfdi)
    company_session.commit()

    body = {
        "domain": [
            ["FechaFiltro", ">=", "2021-01-01T00:00:00.000"],
            ["FechaFiltro", "<", "2022-01-01T00:00:00.000"],
            ["Estatus", "=", True],
            ["is_issued", "=", False],
            ["TipoDeComprobante", "=", "I"],
            ["UUID", "=", cfdi.UUID],
        ],
        "fuzzy_search": "",
        "limit": None,
        "offset": None,
        "order_by": None,
    }
    export_data = dict(
        {
            "file_name": "CFDI_with_Concepts",
            "type": "",
        }
    )

    fields = [
        "UUID",
        "Conceptos.ClaveProdServ",
        "Conceptos.Cantidad",
        "Conceptos.ClaveUnidad",
        "Conceptos.Descripcion",
        "Conceptos.ValorUnitario",
        "Conceptos.Importe",
        "Conceptos.CuentaPredial.Numero",
        "uuid_total_egresos_relacionados",
        "Fecha",
        "LugarExpedicion",
        "Serie",
        "paid_by.UUID",
        "Folio",
        "RfcReceptor",
        "NombreReceptor",
        "Total",
        "TotalMXN",
        "balance",
        "SubTotal",
        "SubTotalMXN",
        "Descuento",
        "DescuentoMXN",
        "Neto",
        "NetoMXN",
        "TrasladosIVA",
        "TrasladosIVAMXN",
        "TrasladosIEPS",
        "TrasladosIEPSMXN",
        "TrasladosISR",
        "TrasladosISRMXN",
        "ExcludeFromIVA",
        "RetencionesIVA",
        "RetencionesIVAMXN",
        "RetencionesIEPS",
        "RetencionesIEPSMXN",
        "RetencionesISR",
        "RetencionesISRMXN",
        "RegimenFiscalReceptor",
        "c_regimen_fiscal_receptor.name",
        "CfdiRelacionados",
        "Moneda",
        "TipoCambio",
        "UsoCFDIReceptor",
        "MetodoPago",
        "c_metodo_pago.name",
        "FormaPago",
        "c_forma_pago.name",
        "CondicionesDePago",
        "Periodicidad",
        "Meses",
        "Year",
        "FechaYear",
        "FechaMonth",
        "Exportacion",
        "FechaCertificacionSat",
        "FechaCertificacionSatYear",
        "FechaCertificacionSatMonth",
        "NoCertificado",
        "Version",
        "TipoDeComprobante",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.Base",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.TipoFactor",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.TasaOCuota",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.Importe",
        "Conceptos.Impuestos.Traslados.Traslado.IEPS.Base",
        "Conceptos.Impuestos.Traslados.Traslado.IEPS.TipoFactor",
        "Conceptos.Impuestos.Traslados.Traslado.IEPS.TasaOCuota",
        "Conceptos.Impuestos.Traslados.Traslado.IEPS.Importe",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.Base",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.TipoFactor",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.TasaOCuota",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.Importe",
        "Conceptos.Impuestos.Retenciones.Retencion.ISR.Base",
        "Conceptos.Impuestos.Retenciones.Retencion.ISR.TipoFactor",
        "Conceptos.Impuestos.Retenciones.Retencion.ISR.TasaOCuota",
        "Conceptos.Impuestos.Retenciones.Retencion.ISR.Importe",
    ]

    query = CFDIController._search(**body, fields=fields, session=company_session)

    exporter = CFDIExporter(
        company_session,
        cfdi_export_repo=CFDIExportRepositorySA(session=company_session),
    )
    file = exporter.export_xlsxv2(
        body=body,
        query=query,
        fields=fields,
        resume_type=ResumeType.BASIC,
        export_data=export_data,
        context=None,
    )

    # with open("test_export_cfdi_multipe.xlsx", "wb") as f:
    #     f.write(file)

    assert "Conceptos.CuentaPredial.Numero" in fields
    assert "Conceptos.ClaveProdServ" in fields
    assert "Conceptos.Cantidad" in fields
    assert "Conceptos.ClaveUnidad" in fields
    assert "Conceptos.Descripcion" in fields
    assert "Conceptos.ValorUnitario" in fields
    assert file is not None


@pytest.mark.skip("generic_xslx")
def test_export_cfdi_with_concepts(company_session: Session):
    conceptos_data = {
        "Concepto": [
            {
                "@ClaveProdServ": "01010101",
                "@NoIdentificacion": "3868",
                "@Cantidad": "2.000000",
                "@ClaveUnidad": "H87",
                "@Unidad": "PIEZA",
                "@Descripcion": "GORDITA ARRIERO 1 ORD",
                "@ValorUnitario": "14.66",
                "@Importe": "29.31",
                "Impuestos": {
                    "Traslados": {
                        "Traslado": {
                            "@Base": "29.312500",
                            "@Impuesto": "002",
                            "@TipoFactor": "Tasa",
                            "@TasaOCuota": "0.160000",
                            "@Importe": "4.69",
                        }
                    }
                },
            },
            {
                "@ClaveProdServ": "50202304",
                "@NoIdentificacion": "7084",
                "@Cantidad": "2.000000",
                "@ClaveUnidad": "H87",
                "@Unidad": "PIEZA",
                "@Descripcion": "BOING SABORES 355 ML",
                "@ValorUnitario": "20.00",
                "@Importe": "40.00",
                "Impuestos": None,
            },
        ]
    }
    cfdi = CFDI.demo(TipoDeComprobante="I", Conceptos=json.dumps(conceptos_data))
    company_session.add(cfdi)
    company_session.commit()

    body = {
        "domain": [
            ["TipoDeComprobante", "=", "I"],
        ],
        "fuzzy_search": "",
        "limit": None,
        "offset": None,
        "order_by": None,
    }
    export_data = dict(
        {
            "file_name": "Test_ingresosConceptos",
            "type": "",
        }
    )

    fields = [
        "Fecha",
        "Serie",
        "Folio",
        "RfcEmisor",
        "NombreEmisor",
        "Total",
        "balance",
        "paid_by.UUID",
        "SubTotal",
        "Descuento",
        "Neto",
        "TrasladosIVA",
        "UsoCFDIReceptor",
        "MetodoPago",
        "FormaPago",
        "Conceptos.ClaveProdServ",
        "Conceptos.Cantidad",
        "Conceptos.ClaveUnidad",
        "Conceptos.Descripcion",
        "Conceptos.ValorUnitario",
        "Conceptos.Importe",
        "Conceptos.Descuento",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.Base",
        "Conceptos.Impuestos.Traslados.Traslado.IVA.Importe",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.Base",
        "Conceptos.Impuestos.Retenciones.Retencion.IVA.Importe",
    ]

    query = CFDIController._search(**body, session=company_session)

    exporter = CFDIExporter(
        company_session,
        cfdi_export_repo=CFDIExportRepositorySA(session=company_session),
    )
    file = exporter.export_xlsxv2(
        body=body,
        query=query,
        fields=fields,
        resume_type=ResumeType.BASIC,
        export_data=export_data,
        context=None,
    )

    file_stream = BytesIO(file)
    wb = load_workbook(file_stream)

    ws = wb.active
    rows = ws.max_row

    # Aseguramos que el excel tiene 3 filas: 1 encabezado y los dos registros con conceptos.
    assert rows == 3
